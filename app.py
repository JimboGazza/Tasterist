# ==========================================================
# TASTERIST — MAIN APPLICATION
# Dashboard-first, stable routing
# ==========================================================

import os
import sys
import sqlite3
import calendar
import re
import subprocess
import json
import html
import tempfile
import secrets
import time
import shutil
import urllib.error
import urllib.request
from collections.abc import Mapping
from functools import wraps
from datetime import date, datetime, timedelta
from pathlib import Path
from urllib.parse import urlsplit

from flask import (
    Flask, g, render_template, request,
    redirect, url_for, flash, session, send_file, abort
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from markupsafe import Markup

import pandas as pd
from openpyxl import load_workbook

# ==========================================================
# APP CONFIG
# ==========================================================

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("TASTERIST_SECRET_KEY", "tasterist-dev-key")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
LOCAL_DB_DIR = os.path.join(DATA_DIR, "db")
LOCAL_DB_FILE = os.path.join(LOCAL_DB_DIR, "tasterist.db")
LOCAL_SHEETS_FALLBACK = os.path.join(DATA_DIR, "taster_sheets")
IMPORT_PREVIEW_DIR = os.path.join(DATA_DIR, "import_previews")
IMPORT_SCRIPT = os.path.join(BASE_DIR, "scripts", "import_taster_sheets.py")
DEFAULT_DB_FILE = (
    "/var/data/tasterist.db"
    if (os.environ.get("RENDER") or os.environ.get("TASTERIST_CANONICAL_HOST"))
    else LOCAL_DB_FILE
)
DB_FILE = os.environ.get("TASTERIST_DB_FILE", DEFAULT_DB_FILE)
IMPORT_LOG_FILE = os.path.join(IMPORT_PREVIEW_DIR, "last_import.log")
IMPORT_META_FILE = os.path.join(IMPORT_PREVIEW_DIR, "last_import_meta.json")
RESTORE_LOG_FILE = os.path.join(IMPORT_PREVIEW_DIR, "last_restore.log")
DAY_ORDER = {
    "Monday": 0,
    "Tuesday": 1,
    "Wednesday": 2,
    "Thursday": 3,
    "Friday": 4,
    "Saturday": 5,
    "Sunday": 6,
}
WEEKDAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]
OWNER_EMAIL = os.environ.get("TASTERIST_OWNER_EMAIL", "james@penninegymnastics.com").strip().lower()
OWNER_NAME = os.environ.get("TASTERIST_OWNER_NAME", "James Gardner").strip() or "James Gardner"
OWNER_RESET_PASSWORD = os.environ.get("TASTERIST_OWNER_RESET_PASSWORD", "").strip()
WEAK_PASSWORDS = {
    "admin123",
    "jammy",
    "password",
    "12345",
    "123456",
    "qwerty",
    "letmein",
}
LOGIN_RATE_LIMIT_WINDOW_SEC = int(os.environ.get("TASTERIST_LOGIN_WINDOW_SEC", "900"))
LOGIN_RATE_LIMIT_ATTEMPTS = int(os.environ.get("TASTERIST_LOGIN_MAX_ATTEMPTS", "8"))
LOGIN_LOCKOUT_SEC = int(os.environ.get("TASTERIST_LOGIN_LOCKOUT_SEC", "900"))
SQLITE_BUSY_TIMEOUT_MS = int(os.environ.get("TASTERIST_SQLITE_BUSY_TIMEOUT_MS", "60000"))
DB_INIT_MAX_RETRIES = int(os.environ.get("TASTERIST_DB_INIT_MAX_RETRIES", "8"))
ADMIN_DAY_PROGRAMMES = ("preschool", "honley", "lockwood")
ADMIN_DAY_HIDDEN_CELLS = {
    ("Monday", "lockwood"),
    ("Tuesday", "preschool"),
    ("Thursday", "preschool"),
    ("Saturday", "preschool"),
}
EMAIL_FROM_DEFAULT = os.environ.get("TASTERIST_EMAIL_FROM", "Tasterist <noreply@tasterist.com>").strip()
DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
DB_BACKEND = os.environ.get("TASTERIST_DB_BACKEND", "").strip().lower()
if DB_BACKEND not in {"sqlite", "postgres"}:
    DB_BACKEND = "postgres" if DATABASE_URL else "sqlite"
if DB_BACKEND == "postgres" and not DATABASE_URL:
    DB_BACKEND = "sqlite"
USING_POSTGRES = DB_BACKEND == "postgres"


def is_env_true(name, default="0"):
    return os.environ.get(name, default).strip().lower() in {"1", "true", "yes", "on"}


def enforce_password_policy():
    # Default off per operator request; can be re-enabled explicitly.
    return is_env_true("TASTERIST_ENFORCE_PASSWORD_POLICY", "0")


def legacy_account_cleanup_enabled():
    # Legacy hardening can remove users; keep disabled unless explicitly requested.
    return is_env_true("TASTERIST_LEGACY_ACCOUNT_CLEANUP", "0")


def should_force_password_change(role, must_change_flag=False, raw_password=None):
    role_name = (role or "").strip().lower()
    if role_name in {"admin", "owner"}:
        return False
    if bool(must_change_flag):
        return True
    if raw_password is not None and enforce_password_policy():
        return bool(password_strength_errors(raw_password))
    return False


def email_owner_only_mode():
    # Hard-safety default: weekly emails only go to owner address.
    return is_env_true("TASTERIST_EMAIL_OWNER_ONLY", "1")


def email_enabled():
    # Explicit opt-in switch to prevent accidental sends on first deploy.
    return is_env_true("TASTERIST_EMAIL_ENABLED", "0")


def safe_internal_target(raw_target):
    target = (raw_target or "").strip()
    if not target:
        return None
    parsed = urlsplit(target)
    if parsed.scheme or parsed.netloc:
        return None
    if not target.startswith("/") or target.startswith("//"):
        return None
    return target


def redact_database_url(raw_url):
    url = (raw_url or "").strip()
    if not url:
        return ""
    parsed = urlsplit(url)
    if not parsed.scheme:
        return url
    netloc = parsed.netloc
    if "@" in netloc:
        userinfo, hostinfo = netloc.rsplit("@", 1)
        if ":" in userinfo:
            user = userinfo.split(":", 1)[0]
            safe_userinfo = f"{user}:***"
        else:
            safe_userinfo = userinfo
        safe_netloc = f"{safe_userinfo}@{hostinfo}"
    else:
        safe_netloc = netloc
    return f"{parsed.scheme}://{safe_netloc}{parsed.path or ''}"


def destructive_imports_enabled():
    # Safety default: replace-all imports stay disabled unless explicitly opted in.
    return is_env_true("TASTERIST_ALLOW_DESTRUCTIVE_IMPORTS", "0")


def _running_in_prod():
    if is_env_true("TASTERIST_FORCE_SECURE_COOKIES", "0"):
        return True
    if os.environ.get("RENDER"):
        return True
    if os.environ.get("TASTERIST_CANONICAL_HOST", "").strip():
        return True
    return False


app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=_running_in_prod(),
    PERMANENT_SESSION_LIFETIME=timedelta(hours=12),
    MAX_CONTENT_LENGTH=100 * 1024 * 1024,
)


def log_runtime_environment():
    python_executable = sys.executable
    active_venv = os.environ.get("VIRTUAL_ENV")

    print(f"🐍 Python executable: {python_executable}")
    if active_venv:
        print(f"📦 Virtual environment: {active_venv}")
    else:
        print("⚠️ Virtual environment: not active")


log_runtime_environment()


def get_import_source_folder():
    configured = os.environ.get("TASTER_SHEETS_FOLDER")
    if configured:
        return os.path.expanduser(configured)

    render_default = "/var/data/taster-sheets"
    if _running_in_prod():
        if USING_POSTGRES and not os.path.isdir("/var/data"):
            return os.path.join("/tmp", "taster-sheets")
        return render_default

    onedrive_default = (
        "/Users/jamesgardner/Library/CloudStorage/OneDrive-Personal/"
        "New Shared Folder/AA Admin/Class Management/Taster Sheets"
    )
    if os.path.isdir(onedrive_default):
        return onedrive_default

    legacy_local = os.path.join(BASE_DIR, "Taster Sheets")
    if os.path.isdir(legacy_local):
        return legacy_local

    return LOCAL_SHEETS_FALLBACK


class RowCompat(Mapping):
    def __init__(self, columns, values):
        self._columns = tuple(columns)
        self._values = tuple(values)
        self._data = {k: v for k, v in zip(self._columns, self._values)}

    def __getitem__(self, key):
        if isinstance(key, int):
            return self._values[key]
        return self._data[key]

    def __iter__(self):
        return iter(self._data)

    def __len__(self):
        return len(self._data)

    def get(self, key, default=None):
        return self._data.get(key, default)


def _replace_qmarks_with_percent_s(sql):
    out = []
    in_single_quote = False
    i = 0
    while i < len(sql):
        ch = sql[i]
        if ch == "'":
            out.append(ch)
            if in_single_quote and i + 1 < len(sql) and sql[i + 1] == "'":
                out.append(sql[i + 1])
                i += 1
            else:
                in_single_quote = not in_single_quote
        elif ch == "?" and not in_single_quote:
            out.append("%s")
        else:
            out.append(ch)
        i += 1
    return "".join(out)


def _translate_sql_for_postgres(sql):
    text = sql
    if text.strip().lower() == "select last_insert_rowid()":
        return "SELECT pg_catalog.lastval()"

    text = re.sub(
        r"CAST\s*\(\s*strftime\('%Y'\s*,\s*([^)]+)\)\s+AS\s+INTEGER\s*\)",
        r"CAST(to_char((\1)::date, 'YYYY') AS INTEGER)",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(
        r"strftime\('%Y-%m'\s*,\s*([^)]+)\)",
        r"to_char((\1)::date, 'YYYY-MM')",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(
        r"strftime\('%w'\s*,\s*([^)]+)\)",
        r"(CAST(EXTRACT(DOW FROM (\1)::date) AS INTEGER)::text)",
        text,
        flags=re.IGNORECASE,
    )
    return _replace_qmarks_with_percent_s(text)


class PostgresCursorCompat:
    def __init__(self, raw_cursor):
        self._cur = raw_cursor

    def _rows_to_compat(self, rows):
        if not rows:
            return []
        cols = [d.name if hasattr(d, "name") else d[0] for d in (self._cur.description or [])]
        return [RowCompat(cols, row) for row in rows]

    def execute(self, sql, args=()):
        translated = _translate_sql_for_postgres(sql)
        self._cur.execute(translated, tuple(args or ()))
        return self

    def executemany(self, sql, seq_of_args):
        translated = _translate_sql_for_postgres(sql)
        payload = [tuple(args or ()) for args in seq_of_args]
        self._cur.executemany(translated, payload)
        return self

    def fetchone(self):
        row = self._cur.fetchone()
        if row is None:
            return None
        cols = [d.name if hasattr(d, "name") else d[0] for d in (self._cur.description or [])]
        return RowCompat(cols, row)

    def fetchall(self):
        return self._rows_to_compat(self._cur.fetchall())

    def fetchmany(self, size=None):
        if size is None:
            rows = self._cur.fetchmany()
        else:
            rows = self._cur.fetchmany(size)
        return self._rows_to_compat(rows)

    def __iter__(self):
        return iter(self.fetchall())

    @property
    def rowcount(self):
        return self._cur.rowcount

    @property
    def description(self):
        return self._cur.description

    def close(self):
        self._cur.close()

    def __getattr__(self, name):
        return getattr(self._cur, name)


class PostgresConnectionCompat:
    def __init__(self, raw_conn):
        self._conn = raw_conn

    def cursor(self):
        return PostgresCursorCompat(self._conn.cursor())

    def execute(self, sql, args=()):
        cur = self.cursor()
        return cur.execute(sql, args)

    def executemany(self, sql, seq_of_args):
        cur = self.cursor()
        return cur.executemany(sql, seq_of_args)

    def commit(self):
        self._conn.commit()

    def rollback(self):
        self._conn.rollback()

    def close(self):
        self._conn.close()

    def __getattr__(self, name):
        return getattr(self._conn, name)


def _connect_postgres():
    import psycopg
    return PostgresConnectionCompat(psycopg.connect(DATABASE_URL))


def _connect_sqlite():
    os.makedirs(os.path.dirname(DB_FILE), exist_ok=True)
    conn = sqlite3.connect(DB_FILE, timeout=max(5, SQLITE_BUSY_TIMEOUT_MS // 1000))
    conn.execute(f"PRAGMA busy_timeout = {SQLITE_BUSY_TIMEOUT_MS}")
    conn.row_factory = sqlite3.Row
    return conn


def open_db_connection():
    if USING_POSTGRES:
        return _connect_postgres()
    return _connect_sqlite()

# ==========================================================
# DATABASE
# ==========================================================

def get_db():
    if "_db" not in g:
        g._db = open_db_connection()
    return g._db

def query(sql, args=()):
    db = get_db()
    cur = db.execute(sql, args)
    rows = cur.fetchall()
    return rows

@app.teardown_appcontext
def close_db(exception):
    db = g.pop("_db", None)
    if db:
        db.close()


def close_request_db_if_open():
    try:
        db = g.pop("_db", None)
    except RuntimeError:
        return
    if db:
        db.close()


@app.after_request
def apply_security_headers(response):
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault("Permissions-Policy", "geolocation=(), microphone=(), camera=()")
    response.headers.setdefault(
        "Content-Security-Policy",
        (
            "default-src 'self'; "
            "img-src 'self' data: https:; "
            "style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
            "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
            "font-src 'self' https://cdn.jsdelivr.net; "
            "connect-src 'self'; frame-ancestors 'self'; base-uri 'self'; form-action 'self'"
        ),
    )
    if _running_in_prod():
        response.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
    return response


def _init_db_once():
    db = open_db_connection()
    cur = db.cursor()

    if USING_POSTGRES:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS tasters (
                id BIGINT PRIMARY KEY,
                child TEXT,
                programme TEXT,
                location TEXT,
                session TEXT,
                class_name TEXT DEFAULT '',
                taster_date DATE,
                notes TEXT,
                attended INTEGER DEFAULT 0,
                club_fees INTEGER DEFAULT 0,
                bg INTEGER DEFAULT 0,
                badge INTEGER DEFAULT 0,
                reschedule_contacted INTEGER DEFAULT 0
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS leavers (
                id BIGINT PRIMARY KEY,
                child TEXT NOT NULL,
                programme TEXT NOT NULL,
                leave_month TEXT NOT NULL,
                leave_date TEXT DEFAULT '',
                class_day TEXT DEFAULT '',
                session TEXT DEFAULT '',
                class_name TEXT DEFAULT '',
                removed_la INTEGER DEFAULT 0,
                removed_bg INTEGER DEFAULT 0,
                added_to_board INTEGER DEFAULT 0,
                reason TEXT DEFAULT '',
                email TEXT DEFAULT '',
                source TEXT DEFAULT 'import'
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS class_sessions (
                id BIGINT PRIMARY KEY,
                programme TEXT NOT NULL,
                location TEXT NOT NULL,
                session_date TEXT NOT NULL DEFAULT '',
                day TEXT NOT NULL,
                class_name TEXT NOT NULL,
                start_time TEXT NOT NULL,
                end_time TEXT NOT NULL DEFAULT '',
                source_file TEXT DEFAULT ''
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id BIGINT PRIMARY KEY,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                full_name TEXT NOT NULL DEFAULT '',
                role TEXT NOT NULL DEFAULT 'staff',
                password_must_change INTEGER NOT NULL DEFAULT 0,
                email_weekly_reports INTEGER NOT NULL DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS user_admin_days (
                id BIGINT PRIMARY KEY,
                user_id BIGINT NOT NULL,
                day_name TEXT NOT NULL,
                programme TEXT NOT NULL,
                UNIQUE(user_id, day_name, programme)
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS audit_logs (
                id BIGINT PRIMARY KEY,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                user_id BIGINT,
                username TEXT NOT NULL DEFAULT '',
                action TEXT NOT NULL,
                entity_type TEXT NOT NULL DEFAULT '',
                entity_id TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'ok',
                details TEXT NOT NULL DEFAULT ''
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS login_attempts (
                ip_key TEXT PRIMARY KEY,
                count INTEGER NOT NULL DEFAULT 0,
                window_start DOUBLE PRECISION NOT NULL DEFAULT 0,
                locked_until DOUBLE PRECISION NOT NULL DEFAULT 0,
                updated_at DOUBLE PRECISION NOT NULL DEFAULT 0
            )
        """)
        cur.execute("ALTER TABLE leavers ADD COLUMN IF NOT EXISTS leave_month TEXT")
        cur.execute("ALTER TABLE leavers ADD COLUMN IF NOT EXISTS leave_date TEXT DEFAULT ''")
        cur.execute("ALTER TABLE leavers ADD COLUMN IF NOT EXISTS class_day TEXT DEFAULT ''")
        cur.execute("ALTER TABLE leavers ADD COLUMN IF NOT EXISTS session TEXT DEFAULT ''")
        cur.execute("ALTER TABLE leavers ADD COLUMN IF NOT EXISTS class_name TEXT DEFAULT ''")
        cur.execute("ALTER TABLE leavers ADD COLUMN IF NOT EXISTS removed_la INTEGER DEFAULT 0")
        cur.execute("ALTER TABLE leavers ADD COLUMN IF NOT EXISTS removed_bg INTEGER DEFAULT 0")
        cur.execute("ALTER TABLE leavers ADD COLUMN IF NOT EXISTS added_to_board INTEGER DEFAULT 0")
        cur.execute("ALTER TABLE leavers ADD COLUMN IF NOT EXISTS reason TEXT DEFAULT ''")
        cur.execute("ALTER TABLE leavers ADD COLUMN IF NOT EXISTS email TEXT DEFAULT ''")
        cur.execute("ALTER TABLE leavers ADD COLUMN IF NOT EXISTS source TEXT DEFAULT 'import'")
        cur.execute("""
            UPDATE leavers
            SET leave_month = substring(leave_date from 1 for 7)
            WHERE (leave_month IS NULL OR trim(leave_month) = '')
              AND leave_date IS NOT NULL
              AND trim(leave_date) <> ''
        """)
        cur.execute("ALTER TABLE class_sessions ADD COLUMN IF NOT EXISTS session_date TEXT NOT NULL DEFAULT ''")
        cur.execute("ALTER TABLE class_sessions ADD COLUMN IF NOT EXISTS source_file TEXT DEFAULT ''")
        cur.execute("ALTER TABLE tasters ADD COLUMN IF NOT EXISTS class_name TEXT DEFAULT ''")
        cur.execute("ALTER TABLE tasters ADD COLUMN IF NOT EXISTS club_fees INTEGER DEFAULT 0")
        cur.execute("ALTER TABLE tasters ADD COLUMN IF NOT EXISTS reschedule_contacted INTEGER DEFAULT 0")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS full_name TEXT NOT NULL DEFAULT ''")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS password_must_change INTEGER NOT NULL DEFAULT 0")
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS email_weekly_reports INTEGER NOT NULL DEFAULT 0")

        for table_name in ("users", "class_sessions", "tasters", "leavers", "user_admin_days", "audit_logs"):
            seq_name = f"{table_name}_id_seq"
            cur.execute(f"CREATE SEQUENCE IF NOT EXISTS {seq_name}")
            max_id_row = cur.execute(f"SELECT COALESCE(MAX(id), 0) FROM {table_name}").fetchone()
            max_id = int(max_id_row[0] or 0)
            if max_id > 0:
                cur.execute(f"SELECT setval('{seq_name}', {max_id}, true)")
            else:
                cur.execute(f"SELECT setval('{seq_name}', 1, false)")
            cur.execute(f"ALTER SEQUENCE {seq_name} OWNED BY {table_name}.id")
            cur.execute(
                f"ALTER TABLE {table_name} ALTER COLUMN id SET DEFAULT nextval('{seq_name}')"
            )
    else:
        cur.execute("PRAGMA journal_mode=WAL")
        cur.execute("""
            CREATE TABLE IF NOT EXISTS tasters (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                child TEXT,
                programme TEXT,
                location TEXT,
                session TEXT,
                class_name TEXT DEFAULT '',
                taster_date DATE,
                notes TEXT,
                attended INTEGER DEFAULT 0,
                club_fees INTEGER DEFAULT 0,
                bg INTEGER DEFAULT 0,
                badge INTEGER DEFAULT 0,
                reschedule_contacted INTEGER DEFAULT 0
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS leavers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                child TEXT NOT NULL,
                programme TEXT NOT NULL,
                leave_month TEXT NOT NULL,
                leave_date TEXT DEFAULT '',
                class_day TEXT DEFAULT '',
                session TEXT DEFAULT '',
                class_name TEXT DEFAULT '',
                removed_la INTEGER DEFAULT 0,
                removed_bg INTEGER DEFAULT 0,
                added_to_board INTEGER DEFAULT 0,
                reason TEXT DEFAULT '',
                email TEXT DEFAULT '',
                source TEXT DEFAULT 'import'
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS class_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                programme TEXT NOT NULL,
                location TEXT NOT NULL,
                session_date TEXT NOT NULL DEFAULT '',
                day TEXT NOT NULL,
                class_name TEXT NOT NULL,
                start_time TEXT NOT NULL,
                end_time TEXT NOT NULL DEFAULT '',
                source_file TEXT DEFAULT ''
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                full_name TEXT NOT NULL DEFAULT '',
                role TEXT NOT NULL DEFAULT 'staff',
                password_must_change INTEGER NOT NULL DEFAULT 0,
                email_weekly_reports INTEGER NOT NULL DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS user_admin_days (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                day_name TEXT NOT NULL,
                programme TEXT NOT NULL,
                UNIQUE(user_id, day_name, programme),
                FOREIGN KEY(user_id) REFERENCES users(id) ON DELETE CASCADE
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL DEFAULT (datetime('now')),
                user_id INTEGER,
                username TEXT NOT NULL DEFAULT '',
                action TEXT NOT NULL,
                entity_type TEXT NOT NULL DEFAULT '',
                entity_id TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT 'ok',
                details TEXT NOT NULL DEFAULT ''
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS login_attempts (
                ip_key TEXT PRIMARY KEY,
                count INTEGER NOT NULL DEFAULT 0,
                window_start REAL NOT NULL DEFAULT 0,
                locked_until REAL NOT NULL DEFAULT 0,
                updated_at REAL NOT NULL DEFAULT 0
            )
        """)
        # Backward-compat: old DBs may still use leave_date only.
        leaver_cols = {
            row[1] for row in cur.execute("PRAGMA table_info(leavers)")
        }
        if "leave_month" not in leaver_cols and "leave_date" in leaver_cols:
            cur.execute("ALTER TABLE leavers ADD COLUMN leave_month TEXT")
            cur.execute("""
                UPDATE leavers
                SET leave_month = substr(leave_date, 1, 7)
                WHERE leave_month IS NULL
            """)
        if "leave_date" not in leaver_cols:
            cur.execute("ALTER TABLE leavers ADD COLUMN leave_date TEXT DEFAULT ''")
        if "class_day" not in leaver_cols:
            cur.execute("ALTER TABLE leavers ADD COLUMN class_day TEXT DEFAULT ''")
        if "session" not in leaver_cols:
            cur.execute("ALTER TABLE leavers ADD COLUMN session TEXT DEFAULT ''")
        if "class_name" not in leaver_cols:
            cur.execute("ALTER TABLE leavers ADD COLUMN class_name TEXT DEFAULT ''")
        if "removed_la" not in leaver_cols:
            cur.execute("ALTER TABLE leavers ADD COLUMN removed_la INTEGER DEFAULT 0")
        if "removed_bg" not in leaver_cols:
            cur.execute("ALTER TABLE leavers ADD COLUMN removed_bg INTEGER DEFAULT 0")
        if "added_to_board" not in leaver_cols:
            cur.execute("ALTER TABLE leavers ADD COLUMN added_to_board INTEGER DEFAULT 0")
        if "reason" not in leaver_cols:
            cur.execute("ALTER TABLE leavers ADD COLUMN reason TEXT DEFAULT ''")
        if "email" not in leaver_cols:
            cur.execute("ALTER TABLE leavers ADD COLUMN email TEXT DEFAULT ''")
        class_cols = {
            row[1] for row in cur.execute("PRAGMA table_info(class_sessions)")
        }
        if "session_date" not in class_cols:
            cur.execute(
                "ALTER TABLE class_sessions ADD COLUMN session_date TEXT NOT NULL DEFAULT ''"
            )
        taster_cols = {
            row[1] for row in cur.execute("PRAGMA table_info(tasters)")
        }
        if "class_name" not in taster_cols:
            cur.execute("ALTER TABLE tasters ADD COLUMN class_name TEXT DEFAULT ''")
        if "club_fees" not in taster_cols:
            cur.execute("ALTER TABLE tasters ADD COLUMN club_fees INTEGER DEFAULT 0")
        if "reschedule_contacted" not in taster_cols:
            cur.execute("ALTER TABLE tasters ADD COLUMN reschedule_contacted INTEGER DEFAULT 0")
        user_cols = {
            row[1] for row in cur.execute("PRAGMA table_info(users)")
        }
        if "full_name" not in user_cols:
            cur.execute("ALTER TABLE users ADD COLUMN full_name TEXT NOT NULL DEFAULT ''")
        if "password_must_change" not in user_cols:
            cur.execute("ALTER TABLE users ADD COLUMN password_must_change INTEGER NOT NULL DEFAULT 0")
        if "email_weekly_reports" not in user_cols:
            cur.execute("ALTER TABLE users ADD COLUMN email_weekly_reports INTEGER NOT NULL DEFAULT 0")

    cur.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS uniq_taster
        ON tasters (child, programme, taster_date, session)
    """)
    cur.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS uniq_leaver
        ON leavers (child, programme, leave_month)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_audit_logs_created_at
        ON audit_logs (created_at DESC)
    """)
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_login_attempts_updated_at
        ON login_attempts (updated_at DESC)
    """)
    cur.execute("DROP INDEX IF EXISTS uniq_class_session")
    cur.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS uniq_class_session
        ON class_sessions (
            programme, session_date, day,
            class_name, start_time, end_time
        )
    """)

    # Keep session format consistent: time-only (e.g. 16:00), no weekday prefix.
    for day_name in (
        "Monday", "Tuesday", "Wednesday", "Thursday",
        "Friday", "Saturday", "Sunday"
    ):
        if USING_POSTGRES:
            cur.execute(
                "UPDATE tasters SET session=trim(substring(session from ?)) WHERE session LIKE ?",
                (len(day_name) + 2, f"{day_name} %"),
            )
        else:
            cur.execute(
                "UPDATE tasters SET session=trim(substr(session, ?)) WHERE session LIKE ?",
                (len(day_name) + 2, f"{day_name} %"),
            )

    users_count = cur.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    owner_bootstrap_password = os.environ.get("TASTERIST_OWNER_BOOTSTRAP_PASSWORD", "").strip()
    if users_count == 0:
        if not owner_bootstrap_password:
            owner_bootstrap_password = secrets.token_urlsafe(16)
            print("⚠️ No users found: created owner with generated bootstrap password.")
            print("⚠️ Set TASTERIST_OWNER_BOOTSTRAP_PASSWORD for predictable first boot credentials.")
        owner_must_change = 0
        cur.execute("""
            INSERT INTO users (username, password_hash, full_name, role, password_must_change, email_weekly_reports)
            VALUES (?, ?, ?, 'owner', ?, 1)
        """, (OWNER_EMAIL, generate_password_hash(owner_bootstrap_password), OWNER_NAME, owner_must_change))

    existing_owner = cur.execute(
        "SELECT id FROM users WHERE lower(username)=?",
        (OWNER_EMAIL,)
    ).fetchone()
    if existing_owner:
        cur.execute("""
            UPDATE users
            SET role='owner',
                full_name=CASE
                    WHEN full_name IS NULL OR trim(full_name)='' THEN ?
                    ELSE full_name
                END
            WHERE id=?
        """, (OWNER_NAME, existing_owner["id"] if isinstance(existing_owner, sqlite3.Row) else existing_owner[0]))
    else:
        if not owner_bootstrap_password:
            owner_bootstrap_password = secrets.token_urlsafe(16)
            print("⚠️ Owner account missing: created owner with generated bootstrap password.")
        owner_must_change = 0
        cur.execute("""
            INSERT INTO users (username, password_hash, full_name, role, password_must_change, email_weekly_reports)
            VALUES (?, ?, ?, 'owner', ?, 1)
        """, (OWNER_EMAIL, generate_password_hash(owner_bootstrap_password), OWNER_NAME, owner_must_change))

    # Break-glass owner reset for cloud recovery.
    # If TASTERIST_OWNER_RESET_PASSWORD is set, owner password is rotated at startup.
    if OWNER_RESET_PASSWORD:
        already_applied = int(cur.execute(
            "SELECT COUNT(*) FROM audit_logs WHERE action='owner_reset_password_env_applied'"
        ).fetchone()[0] or 0) > 0
        if not already_applied or is_env_true("TASTERIST_OWNER_RESET_ALWAYS", "0"):
            cur.execute("""
                UPDATE users
                SET password_hash=?,
                    password_must_change=0,
                    role='owner'
                WHERE lower(username)=?
            """, (generate_password_hash(OWNER_RESET_PASSWORD), OWNER_EMAIL))
            cur.execute(
                """
                INSERT INTO audit_logs (created_at, username, action, entity_type, entity_id, status, details)
                VALUES (?, 'system', 'owner_reset_password_env_applied', 'user', ?, 'ok', ?)
                """,
                (
                    datetime.now().isoformat(timespec="seconds"),
                    OWNER_EMAIL,
                    "Owner password reset applied from TASTERIST_OWNER_RESET_PASSWORD",
                ),
            )
            print("⚠️ Owner password reset applied from TASTERIST_OWNER_RESET_PASSWORD.")

    if legacy_account_cleanup_enabled():
        # Remove insecure historical default admin bootstrap accounts.
        insecure_usernames = {
            "admin",
            os.environ.get("TASTERIST_ADMIN_USER", "admin").strip().lower(),
        }
        for username in insecure_usernames:
            if not username or username == OWNER_EMAIL:
                continue
            cur.execute("DELETE FROM user_admin_days WHERE user_id IN (SELECT id FROM users WHERE lower(username)=?)", (username,))
            cur.execute("DELETE FROM users WHERE lower(username)=?", (username,))

        # Remove non-owner users still using known weak/default passwords.
        user_rows = cur.execute("SELECT id, username, password_hash FROM users").fetchall()
        for row in user_rows:
            row_id = row["id"] if isinstance(row, sqlite3.Row) else row[0]
            row_username = (row["username"] if isinstance(row, sqlite3.Row) else row[1]).strip().lower()
            row_hash = row["password_hash"] if isinstance(row, sqlite3.Row) else row[2]
            if any(check_password_hash(row_hash, weak) for weak in WEAK_PASSWORDS):
                if row_username == OWNER_EMAIL:
                    continue
                cur.execute("DELETE FROM user_admin_days WHERE user_id=?", (row_id,))
                cur.execute("DELETE FROM users WHERE id=?", (row_id,))

    normalise_existing_child_names(db)
    db.commit()
    db.close()


def init_db():
    for attempt in range(1, DB_INIT_MAX_RETRIES + 1):
        try:
            _init_db_once()
            return
        except Exception as exc:
            msg = str(exc).lower()
            retryable_sqlite = (not USING_POSTGRES) and isinstance(exc, sqlite3.OperationalError) and "locked" in msg
            retryable_pg = USING_POSTGRES and ("could not connect" in msg or "connection refused" in msg or "timeout" in msg)
            if (retryable_sqlite or retryable_pg) and attempt < DB_INIT_MAX_RETRIES:
                print(f"⚠️ DB init retry ({attempt}/{DB_INIT_MAX_RETRIES}): {exc}")
                time.sleep(1.5)
                continue
            raise


def maybe_restore_sqlite_from_postgres():
    if USING_POSTGRES:
        return
    auto_raw = os.environ.get("TASTERIST_AUTO_RESTORE_FROM_POSTGRES")
    # Safety-first: explicit opt-in only.
    auto_enabled = False if auto_raw is None else auto_raw.strip().lower() in {"1", "true", "yes", "on"}
    if not auto_enabled:
        return

    postgres_url = os.environ.get("DATABASE_URL", "").strip()
    if not postgres_url:
        return

    try:
        import psycopg
    except Exception as exc:
        print(f"⚠️ Auto-restore skipped: psycopg unavailable ({exc})")
        return

    table_order = (
        "users",
        "class_sessions",
        "tasters",
        "leavers",
        "user_admin_days",
        "audit_logs",
    )

    sqlite_conn = None
    pg_conn = None
    try:
        sqlite_conn = sqlite3.connect(DB_FILE, timeout=max(5, SQLITE_BUSY_TIMEOUT_MS // 1000))
        sqlite_conn.execute(f"PRAGMA busy_timeout = {SQLITE_BUSY_TIMEOUT_MS}")
        sqlite_cur = sqlite_conn.cursor()
        local_tasters = sqlite_cur.execute("SELECT COUNT(*) FROM tasters").fetchone()[0]
        if int(local_tasters or 0) > 0:
            print(f"ℹ️ Auto-restore skipped: SQLite already has tasters ({local_tasters}).")
            return

        pg_conn = psycopg.connect(postgres_url)
        with pg_conn.cursor() as pg_cur:
            pg_cur.execute("SELECT COUNT(*) FROM tasters")
            pg_tasters = int(pg_cur.fetchone()[0] or 0)
        if pg_tasters <= 0:
            print("ℹ️ Auto-restore skipped: Postgres has no taster rows.")
            return

        if os.path.exists(DB_FILE) and os.path.getsize(DB_FILE) > 0:
            ts = datetime.now().strftime("%Y%m%d-%H%M%S")
            backup_file = os.path.join(
                os.path.dirname(DB_FILE),
                f"tasterist-autobackup-before-pg-restore-{ts}.db",
            )
            shutil.copy2(DB_FILE, backup_file)
            print(f"ℹ️ Auto-restore backup created: {backup_file}")

        def sqlite_columns(table_name):
            rows = sqlite_conn.execute(f"PRAGMA table_info({table_name})").fetchall()
            return [r[1] for r in rows]

        def postgres_columns(table_name):
            with pg_conn.cursor() as pg_cur:
                pg_cur.execute(
                    """
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_schema='public' AND table_name=%s
                    ORDER BY ordinal_position
                    """,
                    (table_name,),
                )
                return [r[0] for r in pg_cur.fetchall()]

        total_synced = 0
        for table_name in table_order:
            src_cols = postgres_columns(table_name)
            dst_cols = sqlite_columns(table_name)
            cols = [c for c in src_cols if c in dst_cols]
            if not cols:
                print(f"⚠️ Auto-restore skipped table {table_name}: no shared columns.")
                continue

            col_list = ", ".join(cols)
            with pg_conn.cursor() as pg_cur:
                pg_cur.execute(f"SELECT {col_list} FROM {table_name}")
                rows = pg_cur.fetchall()

            if rows:
                placeholders = ", ".join(["?"] * len(cols))
                if "id" in cols:
                    updates = ", ".join([f"{c}=excluded.{c}" for c in cols if c != "id"])
                    sql = (
                        f"INSERT INTO {table_name} ({col_list}) VALUES ({placeholders}) "
                        f"ON CONFLICT(id) DO UPDATE SET {updates}"
                    )
                else:
                    sql = f"INSERT INTO {table_name} ({col_list}) VALUES ({placeholders})"
                sqlite_cur.executemany(sql, rows)
                sqlite_conn.commit()

            total_synced += len(rows)
            print(f"ℹ️ Auto-restore {table_name}: {len(rows)} row(s)")

        final_count = sqlite_cur.execute("SELECT COUNT(*) FROM tasters").fetchone()[0]
        print(
            "✅ Auto-restore complete from Postgres. "
            f"Total rows synced: {total_synced}; tasters now: {final_count}"
        )
    except Exception as exc:
        print(f"⚠️ Auto-restore from Postgres failed: {exc}")
    finally:
        if sqlite_conn is not None:
            sqlite_conn.close()
        if pg_conn is not None:
            pg_conn.close()


def current_user():
    user_id = session.get("user_id")
    if not user_id:
        return None
    row = query(
        "SELECT id, username, full_name, role, password_must_change, email_weekly_reports FROM users WHERE id=?",
        (user_id,)
    )
    if not row:
        return None
    user = dict(row[0])
    assignments = query("""
        SELECT day_name, programme
        FROM user_admin_days
        WHERE user_id=?
        ORDER BY day_name, programme
    """, (user["id"],))
    user["admin_days"] = [dict(a) for a in assignments]
    return user


def user_initials(text):
    tokens = [t for t in re.split(r"[^A-Za-z0-9]+", text or "") if t]
    if not tokens:
        return "U"
    if len(tokens) == 1:
        return tokens[0][:2].upper()
    return (tokens[0][0] + tokens[1][0]).upper()


def _coerce_temporal_value(value):
    if value is None:
        return ("empty", None)
    if isinstance(value, datetime):
        return ("datetime", value)
    if isinstance(value, date):
        return ("date", value)

    text = str(value).strip()
    if not text:
        return ("empty", None)

    month_match = re.fullmatch(r"(\d{4})-(\d{2})", text)
    if month_match:
        year = int(month_match.group(1))
        month = int(month_match.group(2))
        if 1 <= month <= 12:
            return ("month", (year, month))

    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        try:
            return ("date", date.fromisoformat(text))
        except ValueError:
            return ("text", text)

    normalized = text
    if normalized.endswith("Z"):
        normalized = normalized[:-1] + "+00:00"
    try:
        parsed = datetime.fromisoformat(normalized)
        has_clock = bool(re.search(r"[ T]\d{2}:\d{2}", text))
        return ("datetime", parsed) if has_clock else ("date", parsed.date())
    except ValueError:
        return ("text", text)


@app.template_filter("uk_date")
def format_uk_date(value):
    kind, parsed = _coerce_temporal_value(value)
    if kind == "date":
        return parsed.strftime("%d/%m/%Y")
    if kind == "datetime":
        return parsed.strftime("%d/%m/%Y")
    if kind == "month":
        year, month = parsed
        return date(year, month, 1).strftime("%B %Y")
    if kind == "empty":
        return ""
    return str(parsed)


@app.template_filter("uk_datetime")
def format_uk_datetime(value):
    kind, parsed = _coerce_temporal_value(value)
    if kind == "datetime":
        return parsed.strftime("%d/%m/%Y %H:%M")
    if kind == "date":
        return parsed.strftime("%d/%m/%Y")
    if kind == "month":
        year, month = parsed
        return date(year, month, 1).strftime("%B %Y")
    if kind == "empty":
        return ""
    return str(parsed)


@app.template_filter("uk_month")
def format_uk_month(value):
    kind, parsed = _coerce_temporal_value(value)
    if kind == "datetime":
        return parsed.strftime("%B %Y")
    if kind == "date":
        return parsed.strftime("%B %Y")
    if kind == "month":
        year, month = parsed
        return date(year, month, 1).strftime("%B %Y")
    if kind == "empty":
        return ""
    return str(parsed)


def normalise_child_name(value):
    text = re.sub(r"\s+", " ", str(value or "").strip())
    if not text:
        return ""
    words = []
    for word in text.split(" "):
        parts = re.split(r"([\-'])", word)
        rebuilt = []
        for part in parts:
            if part in {"-", "'"}:
                rebuilt.append(part)
            elif part:
                rebuilt.append(part[:1].upper() + part[1:].lower())
        words.append("".join(rebuilt))
    return " ".join(words)


def normalise_existing_child_names(db):
    changed = 0
    for table in ("tasters", "leavers"):
        rows = db.execute(
            f"SELECT id, child FROM {table} WHERE child IS NOT NULL AND trim(child)<>''"
        ).fetchall()
        for row in rows:
            current = row["child"] if isinstance(row, sqlite3.Row) else row[1]
            updated = normalise_child_name(current)
            if updated and updated != current:
                row_id = row["id"] if isinstance(row, sqlite3.Row) else row[0]
                db.execute(f"UPDATE {table} SET child=? WHERE id=?", (updated, row_id))
                changed += 1
    return changed


def admin_day_cell_allowed(day_name, programme):
    if day_name not in DAY_ORDER:
        return False
    if programme not in ADMIN_DAY_PROGRAMMES:
        return False
    if day_name == "Sunday":
        return False
    return (day_name, programme) not in ADMIN_DAY_HIDDEN_CELLS


def build_admin_day_grouped_options():
    grouped = []
    for day_name in WEEKDAY_NAMES:
        if day_name == "Sunday":
            continue
        cells = []
        for programme in ADMIN_DAY_PROGRAMMES:
            cells.append({
                "programme": programme,
                "value": f"{day_name}|{programme}",
                "visible": admin_day_cell_allowed(day_name, programme),
            })
        grouped.append({
            "day_name": day_name,
            "cells": cells,
        })
    return grouped


def parse_admin_day_values(values):
    allowed_pairs = []
    seen = set()
    for value in values:
        if "|" not in value:
            continue
        day_name, programme = value.split("|", 1)
        if not admin_day_cell_allowed(day_name, programme):
            continue
        key = (day_name, programme)
        if key in seen:
            continue
        seen.add(key)
        allowed_pairs.append(key)
    return allowed_pairs


def is_admin_user(user):
    if not user:
        return False
    return user.get("role") in {"admin", "owner"}


def is_owner_user(user):
    if not user:
        return False
    return user.get("role") == "owner"


def owner_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        user = current_user()
        if not is_owner_user(user):
            flash("Owner access only.", "warning")
            return redirect(url_for("dashboard"))
        return view_func(*args, **kwargs)
    return wrapped


def password_strength_errors(password):
    value = password or ""
    errors = []
    if len(value) < 7:
        errors.append("must be at least 7 characters")
    if not re.search(r"[A-Z]", value):
        errors.append("must include an uppercase letter")
    if not re.search(r"\d", value):
        errors.append("must include a number")
    return errors


def is_password_weak_literal(password):
    return (password or "").strip().lower() in WEAK_PASSWORDS


def get_csrf_token():
    token = session.get("_csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["_csrf_token"] = token
    return token


def csrf_field():
    token = get_csrf_token()
    return Markup(f'<input type="hidden" name="_csrf_token" value="{token}">')


def validate_csrf_token():
    sent = request.form.get("_csrf_token", "")
    expected = session.get("_csrf_token", "")
    if not sent or not expected:
        return False
    return secrets.compare_digest(sent, expected)


def client_ip_key():
    forwarded = request.headers.get("X-Forwarded-For", "")
    if forwarded:
        return forwarded.split(",", 1)[0].strip() or "unknown"
    return request.remote_addr or "unknown"


def is_login_rate_limited(ip_key):
    now = time.time()
    row = get_db().execute(
        "SELECT count, window_start, locked_until FROM login_attempts WHERE ip_key=?",
        (ip_key,),
    ).fetchone()
    if not row:
        return False, 0
    locked_until = float(row["locked_until"] or 0)
    window_start = float(row["window_start"] or 0)
    if locked_until > now:
        wait_sec = int(locked_until - now)
        return True, max(wait_sec, 1)
    if now - window_start > LOGIN_RATE_LIMIT_WINDOW_SEC:
        db = get_db()
        db.execute("DELETE FROM login_attempts WHERE ip_key=?", (ip_key,))
        db.commit()
    return False, 0


def record_failed_login(ip_key):
    db = get_db()
    now = time.time()
    row = db.execute(
        "SELECT count, window_start, locked_until FROM login_attempts WHERE ip_key=?",
        (ip_key,),
    ).fetchone()

    if not row or now - float(row["window_start"] or 0) > LOGIN_RATE_LIMIT_WINDOW_SEC:
        count = 1
        window_start = now
        locked_until = 0
    else:
        count = int(row["count"] or 0) + 1
        window_start = float(row["window_start"] or now)
        locked_until = float(row["locked_until"] or 0)

    if count >= LOGIN_RATE_LIMIT_ATTEMPTS:
        locked_until = now + LOGIN_LOCKOUT_SEC
        count = 0
        window_start = now

    db.execute(
        """
        INSERT INTO login_attempts (ip_key, count, window_start, locked_until, updated_at)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(ip_key) DO UPDATE SET
            count=excluded.count,
            window_start=excluded.window_start,
            locked_until=excluded.locked_until,
            updated_at=excluded.updated_at
        """,
        (ip_key, count, window_start, locked_until, now),
    )
    db.commit()


def clear_login_failures(ip_key):
    db = get_db()
    db.execute("DELETE FROM login_attempts WHERE ip_key=?", (ip_key,))
    db.commit()


def admin_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        user = current_user()
        if not is_admin_user(user):
            flash("Admin access only.", "warning")
            return redirect(url_for("account_settings"))
        return view_func(*args, **kwargs)
    return wrapped


def log_audit(action, entity_type="", entity_id="", details="", status="ok"):
    db = get_db()
    user = current_user()
    user_id = user["id"] if user else None
    username = (user["username"] if user else "system")
    db.execute("""
        INSERT INTO audit_logs (
            created_at, user_id, username, action,
            entity_type, entity_id, status, details
        )
        VALUES (?,?,?,?,?,?,?,?)
    """, (
        datetime.now().isoformat(timespec="seconds"),
        user_id,
        username,
        action,
        entity_type,
        str(entity_id or ""),
        status,
        details[:1000],
    ))
    db.commit()


@app.context_processor
def inject_current_user():
    user = current_user()
    return {
        "current_user": user,
        "is_admin_user": is_admin_user(user),
        "is_owner_user": is_owner_user(user),
        "destructive_imports_enabled": destructive_imports_enabled(),
        "email_owner_only_mode": email_owner_only_mode(),
        "user_initials": user_initials,
        "csrf_token": get_csrf_token,
        "csrf_field": csrf_field,
    }


@app.before_request
def enforce_canonical_host():
    canonical_host = os.environ.get("TASTERIST_CANONICAL_HOST", "").strip().lower()
    if not canonical_host:
        return None
    if request.path == "/health":
        return None
    # Safari can fail large file uploads when POST is redirected between hosts.
    if request.method != "GET":
        return None
    host = request.host.split(":", 1)[0].strip().lower()
    if host in {canonical_host, "localhost", "127.0.0.1"}:
        return None
    full_path = request.full_path if request.query_string else request.path
    # Keep one public host and force HTTPS for staff-facing links/bookmarks.
    return redirect(f"https://{canonical_host}{full_path}", code=301)


@app.before_request
def enforce_csrf():
    if request.method not in {"POST", "PUT", "PATCH", "DELETE"}:
        return None
    if request.endpoint in {"static", "health", "cron_weekly_admin_report"}:
        return None
    if not validate_csrf_token():
        abort(400, description="Invalid CSRF token")
    return None


@app.before_request
def require_login():
    allowed = {"login", "signup", "static", "health", "cron_weekly_admin_report"}
    if request.endpoint in allowed:
        return None
    user = current_user()
    if user is None:
        return redirect(url_for("login", next=request.path))
    must_change = should_force_password_change(
        user.get("role"),
        must_change_flag=(bool(session.get("must_change_password")) or bool(user.get("password_must_change"))),
    )
    if must_change and request.endpoint not in {
        "account_settings", "logout", "static", "health"
    }:
        flash("Security update: set a stronger password to continue.", "warning")
        return redirect(url_for("account_settings"))
    return None


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user() is not None:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        ip_key = client_ip_key()
        limited, wait_sec = is_login_rate_limited(ip_key)
        if limited:
            flash(f"Too many sign-in attempts. Try again in about {wait_sec} seconds.", "danger")
            return render_template("login.html"), 429

        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")
        if not username or not password:
            record_failed_login(ip_key)
            flash("Enter username and password.", "warning")
            return render_template("login.html"), 400

        user_rows = query(
            "SELECT id, username, role, password_hash, password_must_change FROM users WHERE username=?",
            (username,),
        )
        if not user_rows or not check_password_hash(user_rows[0]["password_hash"], password):
            record_failed_login(ip_key)
            flash("Invalid username or password.", "danger")
            return render_template("login.html"), 401

        clear_login_failures(ip_key)
        session["user_id"] = user_rows[0]["id"]
        session.permanent = True
        session["must_change_password"] = should_force_password_change(
            user_rows[0]["role"],
            must_change_flag=user_rows[0]["password_must_change"],
            raw_password=password,
        )
        log_audit("login", entity_type="user", entity_id=user_rows[0]["id"], details="Successful login")
        flash("Signed in.", "success")

        if session.get("must_change_password"):
            flash("Security update: please change your password now.", "warning")
            return redirect(url_for("account_settings"))

        target = safe_internal_target(request.args.get("next")) or url_for("dashboard")
        return redirect(target)

    return render_template("login.html")


@app.route("/signup", methods=["GET", "POST"])
def signup():
    abort(404)


@app.route("/logout", methods=["POST"])
def logout():
    user = current_user()
    if user:
        log_audit("logout", entity_type="user", entity_id=user["id"], details="Signed out")
    session.pop("user_id", None)
    session.pop("must_change_password", None)
    flash("Signed out.", "success")
    return redirect(url_for("login"))


def load_last_import_data():
    if not os.path.exists(IMPORT_LOG_FILE):
        return None

    with open(IMPORT_LOG_FILE, "r", encoding="utf-8") as f:
        log_text = f.read().strip()

    if not log_text:
        return None

    run_at = None
    exit_code = None
    if os.path.exists(IMPORT_META_FILE):
        try:
            with open(IMPORT_META_FILE, "r", encoding="utf-8") as f:
                meta = json.load(f)
            run_at = meta.get("run_at")
            exit_code = meta.get("exit_code")
        except (OSError, json.JSONDecodeError):
            run_at = None
            exit_code = None

    if not run_at:
        run_at = datetime.fromtimestamp(
            os.path.getmtime(IMPORT_LOG_FILE)
        ).isoformat(timespec="seconds")

    file_lines = [
        line for line in log_text.splitlines()
        if line.startswith("📘 FILE:")
    ]
    warning_lines = [
        line for line in log_text.splitlines()
        if "⚠️" in line
    ]

    taster_matches = re.findall(r"✔ Tasters:\s*(\d+)", log_text)
    leaver_matches = re.findall(r"✔ Leavers:\s*(\d+)", log_text)

    try:
        run_at_display = datetime.fromisoformat(run_at).strftime("%d %b %Y %H:%M")
    except ValueError:
        run_at_display = run_at

    return {
        "log_text": log_text,
        "run_at": run_at_display,
        "exit_code": exit_code,
        "files_processed": len(file_lines),
        "warnings": warning_lines,
        "total_tasters": int(taster_matches[-1]) if taster_matches else None,
        "total_leavers": int(leaver_matches[-1]) if leaver_matches else None,
    }


def three_month_cutoff_date(today_dt):
    cutoff_month = today_dt.month - 3
    cutoff_year = today_dt.year
    while cutoff_month <= 0:
        cutoff_month += 12
        cutoff_year -= 1
    cutoff_day = min(today_dt.day, calendar.monthrange(cutoff_year, cutoff_month)[1])
    return date(cutoff_year, cutoff_month, cutoff_day)


def _weekly_report_assignment_set(db, user_id):
    if not user_id:
        return set()
    rows = db.execute(
        "SELECT day_name, programme FROM user_admin_days WHERE user_id=?",
        (user_id,),
    ).fetchall()
    return {(r["day_name"], r["programme"]) for r in rows}


def _row_applies_to_assignments(assignment_set, day_name, programme):
    if not assignment_set:
        return True
    return (day_name, programme) in assignment_set


def weekday_from_iso(value):
    try:
        return datetime.fromisoformat(str(value)).strftime("%A")
    except ValueError:
        return ""


def build_weekly_admin_report_context(db, user_id=None):
    today_dt = date.today()
    today_iso = today_dt.isoformat()
    week_start_dt = today_dt - timedelta(days=6)
    week_start_iso = week_start_dt.isoformat()
    cutoff_dt = three_month_cutoff_date(today_dt)
    cutoff_iso = cutoff_dt.isoformat()
    month_key = today_dt.strftime("%Y-%m")
    assignment_set = _weekly_report_assignment_set(db, user_id)

    raw_followups = db.execute(
        """
        SELECT
            id, child, programme, taster_date, session, class_name,
            attended, club_fees, bg, badge, reschedule_contacted
        FROM tasters
        WHERE taster_date>=?
          AND taster_date<=?
          AND (attended=0 OR club_fees=0 OR bg=0 OR badge=0)
        ORDER BY taster_date DESC, programme, session, child
        """,
        (cutoff_iso, today_iso),
    ).fetchall()

    followups = []
    followups_by_programme = {"preschool": 0, "honley": 0, "lockwood": 0}
    this_week_open = 0
    for row in raw_followups:
        row_dict = dict(row)
        day_name = weekday_from_iso(row_dict["taster_date"])
        if not day_name:
            continue
        if not _row_applies_to_assignments(assignment_set, day_name, row_dict["programme"]):
            continue
        pending = []
        if int(row_dict["attended"] or 0) == 0:
            pending.append("Attended")
        if int(row_dict["club_fees"] or 0) == 0:
            pending.append("Club Fees")
        if int(row_dict["bg"] or 0) == 0:
            pending.append("BG")
        if int(row_dict["badge"] or 0) == 0:
            pending.append("Badge")
        row_dict["pending_labels"] = pending
        row_dict["day_name"] = day_name
        followups.append(row_dict)
        if row_dict["programme"] in followups_by_programme:
            followups_by_programme[row_dict["programme"]] += 1
        if row_dict["taster_date"] >= week_start_iso:
            this_week_open += 1

    raw_members = db.execute(
        """
        SELECT taster_date, programme
        FROM tasters
        WHERE strftime('%Y-%m', taster_date)=?
          AND attended=1
          AND club_fees=1
          AND bg=1
          AND badge=1
        """,
        (month_key,),
    ).fetchall()
    members_month = 0
    for row in raw_members:
        day_name = weekday_from_iso(row["taster_date"])
        if not day_name:
            continue
        if not _row_applies_to_assignments(assignment_set, day_name, row["programme"]):
            continue
        members_month += 1

    raw_leavers = db.execute(
        """
        SELECT programme, leave_date, class_day, session
        FROM leavers
        WHERE leave_month=?
        """,
        (month_key,),
    ).fetchall()
    leavers_month = 0
    for row in raw_leavers:
        day_name = (
            extract_day_name(row["class_day"])
            or extract_day_name(row["session"])
        )
        leave_date = str(row["leave_date"] or "").strip()
        if not day_name and leave_date:
            try:
                day_name = datetime.fromisoformat(leave_date).strftime("%A")
            except ValueError:
                day_name = ""
        if not _row_applies_to_assignments(assignment_set, day_name, row["programme"]):
            continue
        leavers_month += 1

    return {
        "today": today_dt,
        "today_iso": today_iso,
        "week_start_iso": week_start_iso,
        "cutoff_iso": cutoff_iso,
        "month_label": today_dt.strftime("%B %Y"),
        "assignment_count": len(assignment_set),
        "followup_total": len(followups),
        "followups_by_programme": followups_by_programme,
        "this_week_open": this_week_open,
        "members_month": members_month,
        "leavers_month": leavers_month,
        "top_followups": followups[:14],
    }


def build_weekly_admin_report_message(context, recipient_name="Team"):
    followup_lines = []
    for row in context["top_followups"]:
        pending = ", ".join(row["pending_labels"]) if row["pending_labels"] else "No checks pending"
        session_label = (row.get("session") or "").strip() or "Session TBD"
        followup_lines.append(
            f"- {row['taster_date']} | {row['programme'].title()} | {row['child']} | {session_label} | Pending: {pending}"
        )

    if not followup_lines:
        followup_lines = ["- No outstanding follow-ups in scope."]

    subject = f"Tasterist Weekly Admin Report - {context['today'].strftime('%d %b %Y')}"
    scope_label = (
        "All programmes"
        if context["assignment_count"] == 0
        else f"Assigned cells only ({context['assignment_count']})"
    )
    text_body = (
        f"Hello {recipient_name},\n\n"
        "Here is your weekly Tasterist admin summary.\n\n"
        f"Scope: {scope_label}\n"
        f"Open follow-ups (last 3 months): {context['followup_total']}\n"
        f"Open follow-ups created this week: {context['this_week_open']}\n"
        f"Members this month ({context['month_label']}): {context['members_month']}\n"
        f"Leavers this month ({context['month_label']}): {context['leavers_month']}\n"
        f"By programme: Preschool {context['followups_by_programme']['preschool']}, "
        f"Honley {context['followups_by_programme']['honley']}, "
        f"Lockwood {context['followups_by_programme']['lockwood']}\n\n"
        "Top items to action:\n"
        + "\n".join(followup_lines)
        + "\n\nRegards,\nTasterist"
    )

    html_lines = []
    for row in context["top_followups"]:
        pending = ", ".join(row["pending_labels"]) if row["pending_labels"] else "No checks pending"
        session_label = (row.get("session") or "").strip() or "Session TBD"
        html_lines.append(
            "<li>"
            f"{html.escape(str(row['taster_date']))} | "
            f"{html.escape(row['programme'].title())} | "
            f"{html.escape(row['child'])} | "
            f"{html.escape(session_label)} | "
            f"Pending: {html.escape(pending)}"
            "</li>"
        )
    if not html_lines:
        html_lines.append("<li>No outstanding follow-ups in scope.</li>")

    html_body = (
        "<!doctype html><html><body style='font-family:Arial,sans-serif;color:#0f172a'>"
        f"<p>Hello {html.escape(recipient_name)},</p>"
        "<p>Here is your weekly Tasterist admin summary.</p>"
        "<ul>"
        f"<li><strong>Scope:</strong> {html.escape(scope_label)}</li>"
        f"<li><strong>Open follow-ups (last 3 months):</strong> {context['followup_total']}</li>"
        f"<li><strong>Open follow-ups created this week:</strong> {context['this_week_open']}</li>"
        f"<li><strong>Members this month ({html.escape(context['month_label'])}):</strong> {context['members_month']}</li>"
        f"<li><strong>Leavers this month ({html.escape(context['month_label'])}):</strong> {context['leavers_month']}</li>"
        f"<li><strong>By programme:</strong> Preschool {context['followups_by_programme']['preschool']}, "
        f"Honley {context['followups_by_programme']['honley']}, "
        f"Lockwood {context['followups_by_programme']['lockwood']}</li>"
        "</ul>"
        "<p><strong>Top items to action</strong></p>"
        f"<ul>{''.join(html_lines)}</ul>"
        "<p>Regards,<br>Tasterist</p>"
        "</body></html>"
    )
    return subject, text_body, html_body


def send_email_via_cloudflare_webhook(to_email, subject, text_body, html_body):
    webhook_url = os.environ.get("TASTERIST_EMAIL_WEBHOOK_URL", "").strip()
    if not webhook_url:
        raise RuntimeError("TASTERIST_EMAIL_WEBHOOK_URL is not set.")

    payload = {
        "from": EMAIL_FROM_DEFAULT,
        "to": to_email,
        "subject": subject,
        "text": text_body,
        "html": html_body,
    }
    req = urllib.request.Request(
        webhook_url,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    webhook_token = os.environ.get("TASTERIST_EMAIL_WEBHOOK_TOKEN", "").strip()
    if webhook_token:
        req.add_header("Authorization", f"Bearer {webhook_token}")

    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            response_body = resp.read().decode("utf-8", errors="replace")
            status_code = int(resp.getcode() or 0)
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Email webhook HTTP {exc.code}: {detail[:300]}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"Email webhook connection failed: {exc}") from exc

    if status_code not in {200, 201, 202}:
        raise RuntimeError(f"Email webhook returned unexpected status {status_code}: {response_body[:300]}")

    return {"status_code": status_code, "body": response_body[:300]}


def send_weekly_admin_report(trigger="manual"):
    db = get_db()
    owner_only = email_owner_only_mode()
    if not email_enabled():
        return {"sent": 0, "recipients": [], "owner_only": owner_only, "disabled": True}

    recipients = []
    if owner_only:
        recipients = [OWNER_EMAIL]
    else:
        rows = db.execute(
            """
            SELECT username
            FROM users
            WHERE email_weekly_reports=1
            ORDER BY username
            """
        ).fetchall()
        recipients = [r["username"].strip().lower() for r in rows if (r["username"] or "").strip()]

    deduped = []
    seen = set()
    for email_addr in recipients:
        e = email_addr.strip().lower()
        if not e or e in seen:
            continue
        seen.add(e)
        deduped.append(e)
    recipients = deduped
    if not recipients:
        return {"sent": 0, "recipients": [], "owner_only": owner_only}

    sent_count = 0
    for recipient in recipients:
        user_row = db.execute(
            """
            SELECT id, full_name
            FROM users
            WHERE lower(username)=?
            """,
            (recipient,),
        ).fetchone()
        user_id = user_row["id"] if user_row and not owner_only else None
        recipient_name = (
            (user_row["full_name"] or "").strip().split(" ")[0]
            if user_row and (user_row["full_name"] or "").strip()
            else "Team"
        )
        context = build_weekly_admin_report_context(db, user_id=user_id)
        subject, text_body, html_body = build_weekly_admin_report_message(context, recipient_name=recipient_name)
        send_email_via_cloudflare_webhook(recipient, subject, text_body, html_body)
        sent_count += 1
        log_audit(
            "weekly_admin_report_email",
            entity_type="user",
            entity_id=recipient,
            details=f"trigger={trigger} sent_to={recipient} open_followups={context['followup_total']}",
        )

    return {
        "sent": sent_count,
        "recipients": recipients,
        "owner_only": owner_only,
        "disabled": False,
    }


def run_import_process(trigger="manual", replace=False):
    if USING_POSTGRES:
        log_text = (
            "Import runner is disabled while Postgres is the primary runtime database. "
            "Use the migration/sync scripts for controlled imports."
        )
        os.makedirs(os.path.dirname(IMPORT_LOG_FILE), exist_ok=True)
        with open(IMPORT_LOG_FILE, "w", encoding="utf-8") as f:
            f.write(log_text + "\n")
        with open(IMPORT_META_FILE, "w", encoding="utf-8") as f:
            json.dump({
                "run_at": datetime.now().isoformat(timespec="seconds"),
                "exit_code": 2,
                "trigger": trigger,
            }, f)
        return 2, log_text

    if replace and not destructive_imports_enabled():
        print("⚠️ Replace-all import requested but blocked (TASTERIST_ALLOW_DESTRUCTIVE_IMPORTS is off).")
        replace = False
    import_source = get_import_source_folder()
    os.makedirs(import_source, exist_ok=True)
    local_fallback = LOCAL_SHEETS_FALLBACK
    timeout_raw = os.environ.get("TASTERIST_IMPORT_TIMEOUT_SEC", "120").strip()
    try:
        timeout_seconds = max(15, int(timeout_raw))
    except ValueError:
        timeout_seconds = 120
    cmd = [
        sys.executable,
        IMPORT_SCRIPT,
        "--folder", import_source,
        "--db", DB_FILE,
    ]
    if replace:
        cmd.append("--apply")
    if os.path.isdir(local_fallback):
        cmd.extend(["--fallback-folder", local_fallback])
    # Release the request DB handle before subprocess touches the same SQLite file.
    close_request_db_if_open()
    try:
        result = subprocess.run(
            cmd, cwd=BASE_DIR, capture_output=True, text=True, timeout=timeout_seconds
        )
    except subprocess.TimeoutExpired as exc:
        log_parts = []
        stdout_txt = (exc.stdout or "").strip()
        stderr_txt = (exc.stderr or "").strip()
        if stdout_txt:
            log_parts.append(stdout_txt)
        if stderr_txt:
            log_parts.append(stderr_txt)
        log_parts.append(f"Import timed out after {timeout_seconds}s.")
        log_text = "\n\n".join(part for part in log_parts if part).strip() or "(No output captured)"
        os.makedirs(os.path.dirname(IMPORT_LOG_FILE), exist_ok=True)
        with open(IMPORT_LOG_FILE, "w", encoding="utf-8") as f:
            f.write(log_text + "\n")
        with open(IMPORT_META_FILE, "w", encoding="utf-8") as f:
            json.dump({
                "run_at": datetime.now().isoformat(timespec="seconds"),
                "exit_code": 124,
                "trigger": trigger
            }, f)
        return 124, log_text
    except Exception as exc:
        log_text = f"Import execution error: {exc}"
        os.makedirs(os.path.dirname(IMPORT_LOG_FILE), exist_ok=True)
        with open(IMPORT_LOG_FILE, "w", encoding="utf-8") as f:
            f.write(log_text + "\n")
        with open(IMPORT_META_FILE, "w", encoding="utf-8") as f:
            json.dump({
                "run_at": datetime.now().isoformat(timespec="seconds"),
                "exit_code": 125,
                "trigger": trigger
            }, f)
        return 125, log_text

    os.makedirs(os.path.dirname(IMPORT_LOG_FILE), exist_ok=True)
    log_parts = []
    if result.stdout:
        log_parts.append(result.stdout.strip())
    if result.stderr:
        log_parts.append(result.stderr.strip())
    log_text = "\n\n".join(part for part in log_parts if part).strip()
    if not log_text:
        log_text = "(No output captured)"

    with open(IMPORT_LOG_FILE, "w", encoding="utf-8") as f:
        f.write(log_text + "\n")

    with open(IMPORT_META_FILE, "w", encoding="utf-8") as f:
        json.dump({
            "run_at": datetime.now().isoformat(timespec="seconds"),
            "exit_code": result.returncode,
            "trigger": trigger
        }, f)

    return result.returncode, log_text

# ==========================================================
# HELPERS
# ==========================================================

def load_tasters_df(programme=None):
    q = "SELECT * FROM tasters"
    args = []

    if programme:
        q += " WHERE programme=?"
        args.append(programme)

    q += " ORDER BY taster_date, session, child"

    if USING_POSTGRES:
        rows = query(q, tuple(args))
        df = pd.DataFrame([dict(r) for r in rows])
    else:
        db = get_db()
        df = pd.read_sql_query(q, db, params=args)
    if not df.empty:
        df["taster_date"] = pd.to_datetime(df["taster_date"]).dt.date
        if "session" in df.columns:
            df["session"] = df["session"].fillna("")
        if "class_name" in df.columns:
            df["class_name"] = df["class_name"].fillna("")
    return df


def normalise_session_label(value):
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    s = re.sub(
        r"^(monday|tuesday|wednesday|thursday|friday|saturday|sunday)\s+",
        "",
        s,
        flags=re.IGNORECASE,
    )
    m = re.search(r"\b(\d{1,2}):(\d{2})(?::\d{2})?\s*([ap]m)?\b", s, flags=re.IGNORECASE)
    if not m:
        return s
    hour = int(m.group(1))
    minute = int(m.group(2))
    meridiem = (m.group(3) or "").lower()
    if minute > 59:
        return s
    if meridiem:
        if hour < 1 or hour > 12:
            return s
        if meridiem == "am":
            hour = 0 if hour == 12 else hour
        else:
            hour = 12 if hour == 12 else hour + 12
    elif hour > 23:
        return s
    return f"{hour:02d}:{minute:02d}"


def parse_hhmm_like(value):
    text = str(value or "").strip()
    match = re.match(r"^([01]?\d|2[0-3]):([0-5]\d)(?::([0-5]\d))?$", text)
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2))
    second = int(match.group(3)) if match.group(3) is not None else None
    return hour, minute, second


def shift_time_value_to_pm(value):
    parsed = parse_hhmm_like(value)
    if not parsed:
        return None
    hour, minute, second = parsed
    if hour <= 0 or hour >= 12:
        return None
    pm_hour = hour + 12
    if second is None:
        return f"{pm_hour:02d}:{minute:02d}"
    return f"{pm_hour:02d}:{minute:02d}:{second:02d}"


def shift_time_value_late_evening_to_day(value):
    parsed = parse_hhmm_like(value)
    if not parsed:
        return None
    hour, minute, second = parsed
    if hour < 20:
        return None
    day_hour = hour - 12
    if second is None:
        return f"{day_hour:02d}:{minute:02d}"
    return f"{day_hour:02d}:{minute:02d}:{second:02d}"


def run_pm_time_fix(force=False, include_preschool=False):
    conn = open_db_connection()
    cur = conn.cursor()

    # Skip repeated auto-fix passes once an auto run has already been recorded.
    if not force:
        existing = cur.execute(
            "SELECT COUNT(*) FROM audit_logs WHERE action='auto_fix_pm_times'"
        ).fetchone()[0]
        if int(existing or 0) > 0:
            conn.close()
            return {
                "applied": False,
                "reason": "already_ran",
                "suspicious": False,
                "morning_non_preschool": 0,
                "pm_non_preschool": 0,
                "tasters_updated": 0,
                "leavers_updated": 0,
                "class_start_updated": 0,
                "class_end_updated": 0,
            }

    profile_rows = cur.execute(
        """
        SELECT programme, session
        FROM tasters
        WHERE trim(COALESCE(session, ''))<>''
        """
    ).fetchall()
    morning_non_preschool = 0
    pm_non_preschool = 0
    for row in profile_rows:
        programme = str(row["programme"] or "").strip().lower()
        parsed = parse_hhmm_like(row["session"])
        if not parsed or programme == "preschool":
            continue
        hour = parsed[0]
        if 1 <= hour <= 11:
            morning_non_preschool += 1
        elif hour >= 12:
            pm_non_preschool += 1

    suspicious = (
        morning_non_preschool >= 60
        and pm_non_preschool <= max(4, morning_non_preschool // 25)
    )
    if not force and not suspicious:
        conn.close()
        return {
            "applied": False,
            "reason": "not_suspicious",
            "suspicious": False,
            "morning_non_preschool": morning_non_preschool,
            "pm_non_preschool": pm_non_preschool,
            "tasters_updated": 0,
            "leavers_updated": 0,
            "class_start_updated": 0,
            "class_end_updated": 0,
        }

    def can_update_programme(programme):
        if include_preschool:
            return True
        return str(programme or "").strip().lower() != "preschool"

    tasters_updated = 0
    leavers_updated = 0
    class_start_updated = 0
    class_end_updated = 0

    taster_rows = cur.execute("SELECT id, programme, session FROM tasters").fetchall()
    for row in taster_rows:
        if not can_update_programme(row["programme"]):
            continue
        shifted = shift_time_value_to_pm(row["session"])
        if shifted and shifted != (row["session"] or ""):
            cur.execute("UPDATE tasters SET session=? WHERE id=?", (shifted, row["id"]))
            tasters_updated += 1

    leaver_rows = cur.execute("SELECT id, programme, session FROM leavers").fetchall()
    for row in leaver_rows:
        if not can_update_programme(row["programme"]):
            continue
        shifted = shift_time_value_to_pm(row["session"])
        if shifted and shifted != (row["session"] or ""):
            cur.execute("UPDATE leavers SET session=? WHERE id=?", (shifted, row["id"]))
            leavers_updated += 1

    class_rows = cur.execute(
        "SELECT id, programme, start_time, end_time FROM class_sessions"
    ).fetchall()
    for row in class_rows:
        if not can_update_programme(row["programme"]):
            continue
        shifted_start = shift_time_value_to_pm(row["start_time"])
        shifted_end = shift_time_value_to_pm(row["end_time"])
        update_start = shifted_start and shifted_start != (row["start_time"] or "")
        update_end = shifted_end and shifted_end != (row["end_time"] or "")
        if not update_start and not update_end:
            continue
        cur.execute(
            "UPDATE class_sessions SET start_time=?, end_time=? WHERE id=?",
            (
                shifted_start if update_start else row["start_time"],
                shifted_end if update_end else row["end_time"],
                row["id"],
            ),
        )
        if update_start:
            class_start_updated += 1
        if update_end:
            class_end_updated += 1

    total_updates = tasters_updated + leavers_updated + class_start_updated + class_end_updated
    if total_updates > 0:
        action = "manual_fix_pm_times" if force else "auto_fix_pm_times"
        details = (
            f"include_preschool={1 if include_preschool else 0} | "
            f"morning_non_preschool={morning_non_preschool} | "
            f"pm_non_preschool={pm_non_preschool} | "
            f"tasters={tasters_updated} | leavers={leavers_updated} | "
            f"class_start={class_start_updated} | class_end={class_end_updated}"
        )
        cur.execute(
            """
            INSERT INTO audit_logs (created_at, username, action, entity_type, entity_id, status, details)
            VALUES (?, 'system', ?, 'system', 'time-fix', 'ok', ?)
            """,
            (datetime.now().isoformat(timespec="seconds"), action, details),
        )
    conn.commit()
    conn.close()
    return {
        "applied": total_updates > 0,
        "reason": "updated" if total_updates > 0 else "nothing_to_update",
        "suspicious": suspicious,
        "morning_non_preschool": morning_non_preschool,
        "pm_non_preschool": pm_non_preschool,
        "tasters_updated": tasters_updated,
        "leavers_updated": leavers_updated,
        "class_start_updated": class_start_updated,
        "class_end_updated": class_end_updated,
    }


def run_late_night_time_fix(force=False):
    conn = open_db_connection()
    cur = conn.cursor()

    if not force:
        existing = cur.execute(
            "SELECT COUNT(*) FROM audit_logs WHERE action='auto_fix_late_night_times'"
        ).fetchone()[0]
        if int(existing or 0) > 0:
            conn.close()
            return {
                "applied": False,
                "reason": "already_ran",
                "suspicious_count": 0,
                "tasters_updated": 0,
                "leavers_updated": 0,
                "class_start_updated": 0,
                "class_end_updated": 0,
            }

    suspicious_count = int(cur.execute("""
        SELECT COUNT(*)
        FROM tasters
        WHERE lower(trim(programme)) IN ('lockwood', 'honley')
          AND trim(COALESCE(session, '')) <> ''
          AND CAST(substr(session, 1, 2) AS INTEGER) >= 20
    """).fetchone()[0] or 0)
    if not force and suspicious_count < 20:
        conn.close()
        return {
            "applied": False,
            "reason": "not_suspicious",
            "suspicious_count": suspicious_count,
            "tasters_updated": 0,
            "leavers_updated": 0,
            "class_start_updated": 0,
            "class_end_updated": 0,
        }

    tasters_updated = 0
    leavers_updated = 0
    class_start_updated = 0
    class_end_updated = 0

    taster_rows = cur.execute("""
        SELECT id, session
        FROM tasters
        WHERE lower(trim(programme)) IN ('lockwood', 'honley')
    """).fetchall()
    for row in taster_rows:
        shifted = shift_time_value_late_evening_to_day(row["session"])
        if shifted and shifted != (row["session"] or ""):
            cur.execute("UPDATE tasters SET session=? WHERE id=?", (shifted, row["id"]))
            tasters_updated += 1

    leaver_rows = cur.execute("""
        SELECT id, session
        FROM leavers
        WHERE lower(trim(programme)) IN ('lockwood', 'honley')
    """).fetchall()
    for row in leaver_rows:
        shifted = shift_time_value_late_evening_to_day(row["session"])
        if shifted and shifted != (row["session"] or ""):
            cur.execute("UPDATE leavers SET session=? WHERE id=?", (shifted, row["id"]))
            leavers_updated += 1

    class_rows = cur.execute("""
        SELECT id, start_time, end_time
        FROM class_sessions
        WHERE lower(trim(programme)) IN ('lockwood', 'honley')
          AND day='Saturday'
    """).fetchall()
    for row in class_rows:
        shifted_start = shift_time_value_late_evening_to_day(row["start_time"])
        shifted_end = shift_time_value_late_evening_to_day(row["end_time"])
        update_start = shifted_start and shifted_start != (row["start_time"] or "")
        update_end = shifted_end and shifted_end != (row["end_time"] or "")
        if not update_start and not update_end:
            continue
        cur.execute(
            "UPDATE class_sessions SET start_time=?, end_time=? WHERE id=?",
            (
                shifted_start if update_start else row["start_time"],
                shifted_end if update_end else row["end_time"],
                row["id"],
            ),
        )
        if update_start:
            class_start_updated += 1
        if update_end:
            class_end_updated += 1

    total_updates = tasters_updated + leavers_updated + class_start_updated + class_end_updated
    if total_updates > 0:
        action = "manual_fix_late_night_times" if force else "auto_fix_late_night_times"
        details = (
            f"suspicious_count={suspicious_count} | "
            f"tasters={tasters_updated} | leavers={leavers_updated} | "
            f"class_start={class_start_updated} | class_end={class_end_updated}"
        )
        cur.execute(
            """
            INSERT INTO audit_logs (created_at, username, action, entity_type, entity_id, status, details)
            VALUES (?, 'system', ?, 'system', 'time-fix', 'ok', ?)
            """,
            (datetime.now().isoformat(timespec="seconds"), action, details),
        )
    conn.commit()
    conn.close()
    return {
        "applied": total_updates > 0,
        "reason": "updated" if total_updates > 0 else "nothing_to_update",
        "suspicious_count": suspicious_count,
        "tasters_updated": tasters_updated,
        "leavers_updated": leavers_updated,
        "class_start_updated": class_start_updated,
        "class_end_updated": class_end_updated,
    }


def maybe_auto_fix_pm_times():
    default_enabled = "0" if USING_POSTGRES else "1"
    if not is_env_true("TASTERIST_AUTO_FIX_PM_TIMES", default_enabled):
        return
    try:
        result = run_pm_time_fix(force=False, include_preschool=False)
        if result.get("applied"):
            print(
                "⏰ Auto PM-time fix applied: "
                f"tasters={result['tasters_updated']}, leavers={result['leavers_updated']}, "
                f"class_start={result['class_start_updated']}, class_end={result['class_end_updated']}"
            )
    except Exception as exc:
        print(f"⚠️ Auto PM-time fix failed: {exc}")


def maybe_auto_fix_late_night_times():
    default_enabled = "0" if USING_POSTGRES else "1"
    if not is_env_true("TASTERIST_AUTO_FIX_LATE_NIGHT_TIMES", default_enabled):
        return
    try:
        result = run_late_night_time_fix(force=False)
        if result.get("applied"):
            print(
                "🌙 Auto late-night time fix applied: "
                f"tasters={result['tasters_updated']}, leavers={result['leavers_updated']}, "
                f"class_start={result['class_start_updated']}, class_end={result['class_end_updated']}"
            )
    except Exception as exc:
        print(f"⚠️ Auto late-night time fix failed: {exc}")


def _hhmm_to_minutes(value):
    s = str(value or "").strip()
    m = re.match(r"^(\d{1,2}):(\d{2})$", s)
    if not m:
        return None
    return int(m.group(1)) * 60 + int(m.group(2))


def infer_class_type(class_name, start_time="", end_time=""):
    name = str(class_name or "").strip().lower()
    if "parkour" in name:
        return "Parkour"

    start_m = _hhmm_to_minutes(start_time)
    end_m = _hhmm_to_minutes(end_time)
    if start_m is not None and end_m is not None and end_m > start_m:
        duration = end_m - start_m
        if duration == 45:
            return "45 Minute Gymnastics"
        if duration == 60:
            return "1 Hour Gymnastics"
        if duration == 90:
            return "1.5 Hour Gymnastics"

    if any(token in name for token in ("1.5", "90 min", "90min", "1h30", "1hr30")):
        return "1.5 Hour Gymnastics"
    if any(token in name for token in ("1 hour", "1hr", "60 min", "60min")):
        return "1 Hour Gymnastics"
    if any(token in name for token in ("45 min", "45min", "45-minute", "45 minute")):
        return "45 Minute Gymnastics"
    if any(token in name for token in ("preschool", "pre-school")):
        return "Pre-School"
    return "Gymnastics"


def extract_day_name(value):
    text = str(value or "").strip().lower()
    if not text:
        return ""
    for day_name in WEEKDAY_NAMES:
        if re.search(rf"\b{day_name.lower()}\b", text):
            return day_name
    return ""


def _parse_iso_date(value):
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def _programme_has_session_templates_for_day(db, programme, target_date):
    if not target_date:
        return False
    day_name = target_date.strftime("%A")
    day_iso = target_date.isoformat()
    row = db.execute(
        """
        SELECT COUNT(*) AS c
        FROM class_sessions
        WHERE programme=?
          AND (session_date=? OR day=?)
        """,
        (programme, day_iso, day_name),
    ).fetchone()
    return int(row["c"] or 0) > 0


def _validate_programme_date_guardrails(db, programme, target_date):
    # Keep this dynamic: only block when there are no matching timetable templates.
    # If Saturday preschool classes are added later, entries will be accepted.
    if programme == "preschool" and target_date.strftime("%A") == "Saturday":
        if not _programme_has_session_templates_for_day(db, programme, target_date):
            return False, "No preschool sessions are configured for Saturdays."
    return True, ""


def _programme_tokens(programme):
    p = (programme or "").lower()
    if p == "preschool":
        return ["preschool", "pre-school"]
    if p == "honley":
        return ["honley"]
    return ["lockwood"]


def _candidate_workbooks(root_path, programme, year):
    if not root_path.exists():
        return []
    tokens = _programme_tokens(programme)
    out = []
    for file_path in sorted(root_path.rglob("*.xlsx")):
        name = file_path.name.lower()
        if name.startswith("~$"):
            continue
        if "taster" not in name or "leaver" not in name:
            continue
        if not any(tok in name for tok in tokens):
            continue
        folder_has_year = str(year) in file_path.as_posix()
        name_has_year = str(year) in name
        score = 0
        if folder_has_year:
            score += 2
        if name_has_year:
            score += 2
        if "tasters and leavers" in name:
            score += 1
        out.append((score, file_path))
    out.sort(key=lambda t: (-t[0], str(t[1]).lower()))
    return [p for _, p in out]


def find_programme_workbook(programme, year, prefer_local=True, local_only=False):
    source_root = Path(get_import_source_folder()).expanduser().resolve()
    fallback_root = Path(LOCAL_SHEETS_FALLBACK).resolve()
    if local_only:
        roots = (fallback_root,)
    else:
        roots = (fallback_root, source_root) if prefer_local else (source_root, fallback_root)
    for root in roots:
        matches = _candidate_workbooks(root, programme, year)
        if matches:
            return matches[0]
    return None


def excel_sync_local_only_mode():
    # Default behaviour:
    # - local/dev: write only to local fallback workbook copies
    # - cloud/render: write to configured TASTER_SHEETS_FOLDER
    raw = os.environ.get("TASTERIST_EXCEL_SYNC_LOCAL_ONLY")
    if raw is not None:
        return raw.strip().lower() in {"1", "true", "yes", "on"}
    return not _running_in_prod()


def _extract_time(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if hasattr(value, "hour") and hasattr(value, "minute"):
        try:
            return f"{int(value.hour):02d}:{int(value.minute):02d}"
        except Exception:
            pass
    s = str(value).strip()
    m = re.search(r"(\d{1,2}):(\d{2})", s)
    if not m:
        return ""
    return f"{int(m.group(1)):02d}:{m.group(2)}"


def _time_matches(target, observed):
    t = _extract_time(target)
    o = _extract_time(observed)
    if not t or not o:
        return False
    if t == o:
        return True
    try:
        t_h, t_m = t.split(":", 1)
        o_h, o_m = o.split(":", 1)
        if t_m != o_m:
            return False
        t_i = int(t_h)
        o_i = int(o_h)
        return (t_i + 12) % 24 == o_i or (o_i + 12) % 24 == t_i
    except ValueError:
        return False


def _parse_sheet_date(value, month_name, year):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        try:
            return date(value.year, value.month, value.day)
        except ValueError:
            return None
    s = str(value).strip().lower()
    if not s:
        return None
    s = re.sub(r"(st|nd|rd|th)", "", s)
    s = re.sub(r"\bof\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m", "%d-%b", "%d %b", "%d %B"):
        try:
            d = datetime.strptime(s, fmt)
            if d.year == 1900:
                d = d.replace(year=year)
            return d.date()
        except ValueError:
            continue
    try:
        return datetime.strptime(f"{s} {month_name} {year}", "%d %B %Y").date()
    except ValueError:
        pass
    try:
        return datetime.strptime(f"{s} {month_name[:3]} {year}", "%d %b %Y").date()
    except ValueError:
        return None


def _find_name_columns_ws(ws, max_scan_rows=25):
    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        cols = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == "name":
                cols.append(c)
        if cols:
            return r, cols
    return None, []


def _find_section_rows_ws(ws, marker):
    hits = []
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == marker.lower():
                hits.append(r)
                break
    return hits


def _find_leaver_header_row_ws(ws, start_row):
    scan_to = min(start_row + 18, ws.max_row)
    for r in range(start_row, scan_to + 1):
        name_cols = []
        has_leave = False
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            s = v.strip().lower()
            if s == "name":
                name_cols.append(c)
            if "leave" in s:
                has_leave = True
        if name_cols and has_leave:
            return r, name_cols
    return None, []


def _build_column_map(ws, name_header_row, name_cols):
    def header_text(col_idx):
        if col_idx < 1 or col_idx > ws.max_column:
            return ""
        v = ws.cell(name_header_row, col_idx).value
        return str(v).strip().lower() if v is not None else ""

    def find_col(name_col, fallback_offset, matcher):
        fallback = name_col + fallback_offset
        for c in range(name_col + 1, min(name_col + 11, ws.max_column + 1)):
            if matcher(header_text(c)):
                return c
        return fallback

    out = {}
    for col in name_cols:
        out[col] = {
            "day_col": col - 1,
            "date_col": find_col(
                col, 1,
                lambda t: "date" in t and ("taster" in t or "date of" in t)
            ),
            "attended_col": find_col(
                col, 2,
                lambda t: ("attended" in t) or ("attend" in t)
            ),
            "club_fees_col": find_col(
                col, 3,
                lambda t: ("paid club fees" in t) or ("club fees" in t) or ("dd" in t and "paid" in t)
            ),
            "bg_col": find_col(
                col, 4,
                lambda t: ("paid bg" in t) or (t == "bg") or ("paid" in t and "bg" in t)
            ),
            "badge_col": find_col(
                col, 5,
                lambda t: ("added bg" in t) or ("badge" in t) or ("account" in t and "bg" in t)
            ),
            "notes_col": find_col(
                col, 6,
                lambda t: ("note" in t) or ("medical" in t)
            ),
            "added_by_col": find_col(
                col, 7,
                lambda t: ("added by" in t) or (t.strip() == "added")
            ),
        }
    return out


def _build_leaver_column_map(ws, header_row, name_cols):
    def header_text(col_idx):
        if col_idx < 1 or col_idx > ws.max_column:
            return ""
        v = ws.cell(header_row, col_idx).value
        return str(v).strip().lower() if v is not None else ""

    def find_col(name_col, fallback_offset, matcher):
        fallback = name_col + fallback_offset
        for c in range(name_col + 1, min(name_col + 11, ws.max_column + 1)):
            if matcher(header_text(c)):
                return c
        return fallback

    out = {}
    for col in name_cols:
        out[col] = {
            "day_col": col - 1,
            "date_col": find_col(
                col, 1,
                lambda t: "date" in t and ("leave" in t or "email" in t)
            ),
            "removed_la_col": find_col(
                col, 2,
                lambda t: ("removed from la" in t) or ("inactive" in t) or ("removed" in t and "la" in t)
            ),
            "removed_bg_col": find_col(
                col, 3,
                lambda t: ("removed from bg" in t) or ("removed" in t and "bg" in t)
            ),
            "board_col": find_col(
                col, 4,
                lambda t: ("leavers board" in t) or ("added" in t and "board" in t)
            ),
            "reason_col": find_col(
                col, 5,
                lambda t: ("reason" in t)
            ),
            "added_by_col": find_col(
                col, 6,
                lambda t: ("added by" in t) or (t.strip() == "added")
            ),
        }
    return out


def _sync_yes_cell(value):
    return "yes" if int(value or 0) == 1 else ""


def sync_taster_to_excel(taster_row, mode="add", changed_field="", actor_initials=""):
    try:
        row = dict(taster_row)
        row_date = datetime.fromisoformat(str(row["taster_date"])).date()
    except Exception:
        return False, "Invalid taster date"

    local_only_sync = excel_sync_local_only_mode()
    workbook = find_programme_workbook(
        row.get("programme"), row_date.year, prefer_local=True, local_only=local_only_sync
    )
    if workbook is None:
        if local_only_sync:
            return False, "Local workbook not found in fallback sheets folder for programme/year"
        return False, "Workbook not found in configured sheets folder for programme/year"

    try:
        wb = load_workbook(workbook)
    except Exception as exc:
        return False, f"Could not open workbook: {exc}"

    sheet_name = MONTH_NAMES[row_date.month - 1]
    if sheet_name not in wb.sheetnames:
        return False, f"Month sheet not found: {sheet_name}"
    ws = wb[sheet_name]

    name_header_row, name_cols = _find_name_columns_ws(ws)
    if not name_cols:
        return False, "No Name columns found"
    leaver_markers = _find_section_rows_ws(ws, "LEAVERS")
    taster_end_row = min(leaver_markers) - 1 if leaver_markers else ws.max_row
    column_map = _build_column_map(ws, name_header_row, name_cols)

    target_day = row_date.strftime("%A")
    target_time = _extract_time(row.get("session"))
    month_name = MONTH_NAMES[row_date.month - 1]
    block_state = {
        col: {"day": "", "date": "", "time": ""}
        for col in name_cols
    }
    target_slot = None
    exact_empty_slot = None
    same_day_slot = None
    any_empty_slot = None

    for r in range(1, taster_end_row + 1):
        for col in name_cols:
            cols = column_map[col]
            day_val = ws.cell(r, cols["day_col"]).value if cols["day_col"] >= 1 else ""
            day_txt = str(day_val).strip() if day_val is not None else ""
            parsed_date = _parse_sheet_date(
                ws.cell(r, cols["date_col"]).value if cols["date_col"] <= ws.max_column else "",
                month_name,
                row_date.year
            )

            if day_txt in WEEKDAY_NAMES:
                block_state[col]["day"] = day_txt
            parsed_time = _extract_time(day_txt)
            if parsed_time:
                block_state[col]["time"] = parsed_time
            if parsed_date:
                block_state[col]["date"] = parsed_date.isoformat()

            if r <= name_header_row:
                continue

            name_val = ws.cell(r, col).value
            name_txt = str(name_val).strip() if name_val is not None else ""
            same_day = block_state[col]["day"] == target_day
            same_time = _time_matches(target_time, block_state[col]["time"]) if target_time else same_day

            if name_txt and name_txt.lower() == str(row["child"]).strip().lower():
                row_date_cell = _parse_sheet_date(
                    ws.cell(r, cols["date_col"]).value if cols["date_col"] <= ws.max_column else "",
                    month_name,
                    row_date.year
                )
                if (row_date_cell == row_date) or (same_day and same_time):
                    target_slot = (r, col, cols)
                    break

            if not name_txt:
                if same_day and same_time and exact_empty_slot is None:
                    exact_empty_slot = (r, col, cols)
                if same_day and same_day_slot is None:
                    same_day_slot = (r, col, cols)
                if any_empty_slot is None:
                    any_empty_slot = (r, col, cols)
        if target_slot:
            break

    if target_slot is None:
        if target_time:
            target_slot = exact_empty_slot
        else:
            target_slot = same_day_slot or exact_empty_slot
    if target_slot is None and not target_time:
        target_slot = any_empty_slot
    if target_slot is None:
        return False, "No writable time-matched slot found on sheet"

    row_idx, name_col, cols = target_slot
    if mode == "add":
        ws.cell(row_idx, name_col).value = row["child"]
        if cols["date_col"] <= ws.max_column:
            ws.cell(row_idx, cols["date_col"]).value = f"{row_date.day} {row_date.strftime('%b')}"
        if cols["notes_col"] <= ws.max_column:
            new_notes = str(row.get("notes") or "").strip()
            if new_notes:
                ws.cell(row_idx, cols["notes_col"]).value = new_notes
        if cols["added_by_col"] <= ws.max_column and actor_initials:
            ws.cell(row_idx, cols["added_by_col"]).value = actor_initials
        if cols["attended_col"] <= ws.max_column:
            ws.cell(row_idx, cols["attended_col"]).value = _sync_yes_cell(row.get("attended", 0))
        if cols["club_fees_col"] <= ws.max_column:
            ws.cell(row_idx, cols["club_fees_col"]).value = _sync_yes_cell(row.get("club_fees", 0))
        if cols["bg_col"] <= ws.max_column:
            ws.cell(row_idx, cols["bg_col"]).value = _sync_yes_cell(row.get("bg", 0))
        if cols["badge_col"] <= ws.max_column:
            ws.cell(row_idx, cols["badge_col"]).value = _sync_yes_cell(row.get("badge", 0))
    elif mode == "status":
        field_col_lookup = {
            "attended": cols["attended_col"],
            "club_fees": cols["club_fees_col"],
            "bg": cols["bg_col"],
            "badge": cols["badge_col"],
        }
        target_col = field_col_lookup.get(changed_field)
        if not target_col or target_col > ws.max_column:
            return False, f"Status column not found for {changed_field}"
        ws.cell(row_idx, target_col).value = _sync_yes_cell(row.get(changed_field, 0))
    elif mode == "contacted":
        if cols["notes_col"] <= ws.max_column and int(row.get("reschedule_contacted", 0) or 0) == 1:
            old_note = str(ws.cell(row_idx, cols["notes_col"]).value or "").strip()
            if "contacted" not in old_note.lower():
                ws.cell(row_idx, cols["notes_col"]).value = (
                    f"{old_note} | Contacted for reschedule" if old_note else "Contacted for reschedule"
                )
    else:
        return False, "Unknown sync mode"

    try:
        wb.save(workbook)
    except Exception as exc:
        return False, f"Could not save workbook: {exc}"
    return True, f"Synced to {workbook.name} ({sheet_name})"


def sync_leaver_to_excel(leaver_row, actor_initials=""):
    try:
        row = dict(leaver_row)
        leave_date_raw = (row.get("leave_date") or "").strip()
        if not leave_date_raw:
            return False, "Leave date missing"
        leave_dt = datetime.fromisoformat(leave_date_raw).date()
    except Exception:
        return False, "Invalid leave date"

    local_only_sync = excel_sync_local_only_mode()
    workbook = find_programme_workbook(
        row.get("programme"), leave_dt.year, prefer_local=True, local_only=local_only_sync
    )
    if workbook is None:
        if local_only_sync:
            return False, "Local workbook not found in fallback sheets folder for programme/year"
        return False, "Workbook not found in configured sheets folder for programme/year"
    try:
        wb = load_workbook(workbook)
    except Exception as exc:
        return False, f"Could not open workbook: {exc}"

    sheet_name = MONTH_NAMES[leave_dt.month - 1]
    if sheet_name not in wb.sheetnames:
        return False, f"Month sheet not found: {sheet_name}"
    ws = wb[sheet_name]

    leaver_markers = _find_section_rows_ws(ws, "LEAVERS")
    if not leaver_markers:
        return False, "LEAVERS section not found"
    header_row, name_cols = _find_leaver_header_row_ws(ws, min(leaver_markers))
    if not header_row or not name_cols:
        return False, "Leaver columns not found"
    column_map = _build_leaver_column_map(ws, header_row, name_cols)
    leaver_start_row = min(leaver_markers)

    target_day = extract_day_name(row.get("class_day")) or leave_dt.strftime("%A")
    target_time = _extract_time(row.get("session"))

    block_state = {
        col: {"day": "", "time": ""}
        for col in name_cols
    }
    target_slot = None
    exact_empty_slot = None
    same_day_slot = None
    same_time_slot = None

    for r in range(leaver_start_row, ws.max_row + 1):
        for col in name_cols:
            cols = column_map[col]
            day_col = cols["day_col"]
            day_val = ws.cell(r, day_col).value if day_col >= 1 else ""
            day_txt = str(day_val).strip() if day_val is not None else ""
            if day_txt in WEEKDAY_NAMES:
                block_state[col]["day"] = day_txt
            parsed_time = _extract_time(day_txt)
            if parsed_time:
                block_state[col]["time"] = parsed_time

            if r <= header_row:
                continue

            name_val = ws.cell(r, col).value
            name_txt = str(name_val).strip() if name_val is not None else ""
            same_day = block_state[col]["day"] == target_day if target_day else True
            same_time = _time_matches(target_time, block_state[col]["time"]) if target_time else True

            if name_txt and name_txt.lower() == str(row.get("child", "")).strip().lower():
                if same_day and (same_time or not target_time):
                    target_slot = (r, col, cols)
                    break

            if not name_txt:
                if same_day and same_time and exact_empty_slot is None:
                    exact_empty_slot = (r, col, cols)
                if same_day and same_day_slot is None:
                    same_day_slot = (r, col, cols)
                if same_time and same_time_slot is None:
                    same_time_slot = (r, col, cols)
        if target_slot:
            break

    if target_slot is None:
        target_slot = exact_empty_slot or same_day_slot or (same_time_slot if not target_day else None)
    if target_slot is None:
        return False, "No writable leaver slot found for selected day/time"

    row_idx, name_col, cols = target_slot
    ws.cell(row_idx, name_col).value = row.get("child", "")
    if cols["date_col"] <= ws.max_column:
        ws.cell(row_idx, cols["date_col"]).value = leave_dt.strftime("%d %b")
    if cols["removed_la_col"] <= ws.max_column:
        ws.cell(row_idx, cols["removed_la_col"]).value = _sync_yes_cell(row.get("removed_la", 0))
    if cols["removed_bg_col"] <= ws.max_column:
        ws.cell(row_idx, cols["removed_bg_col"]).value = _sync_yes_cell(row.get("removed_bg", 0))
    if cols["board_col"] <= ws.max_column:
        ws.cell(row_idx, cols["board_col"]).value = _sync_yes_cell(row.get("added_to_board", 0))
    if cols["reason_col"] <= ws.max_column:
        reason_txt = str(row.get("reason") or "").strip()
        if reason_txt:
            ws.cell(row_idx, cols["reason_col"]).value = reason_txt
    if cols["added_by_col"] <= ws.max_column and actor_initials:
        ws.cell(row_idx, cols["added_by_col"]).value = actor_initials

    try:
        wb.save(workbook)
    except Exception as exc:
        return False, f"Could not save workbook: {exc}"
    return True, f"Synced to {workbook.name} ({sheet_name})"


def get_day_programme_options():
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    programmes = ["preschool", "honley", "lockwood"]
    options = []
    for d in days:
        for p in programmes:
            options.append({
                "value": f"{d}|{p}",
                "day_name": d,
                "programme": p,
                "label": f"{d} • {p.title()}",
            })
    return options


def build_week_schedule(programme, week_start):
    week_days = []
    db = get_db()
    today = date.today()
    window_start = today.isoformat()
    window_end = (today + timedelta(days=30)).isoformat()
    programme_has_templates = int(db.execute(
        "SELECT COUNT(*) AS c FROM class_sessions WHERE programme=?",
        (programme,),
    ).fetchone()["c"] or 0) > 0

    for offset in range(7):
        day_date = week_start + timedelta(days=offset)
        day_str = day_date.isoformat()
        day_name = day_date.strftime("%A")

        rows = db.execute("""
            SELECT class_name, start_time, end_time, location
            FROM class_sessions
            WHERE programme=? AND session_date=?
            ORDER BY start_time, class_name
        """, (programme, day_str)).fetchall()

        source_mode = "dated"
        if not rows:
            rows = db.execute("""
                SELECT class_name, start_time, end_time, location
                FROM class_sessions
                WHERE programme=? AND day=?
                ORDER BY start_time, class_name
            """, (programme, day_name)).fetchall()
            source_mode = "weekly"

        # Only derive from taster history when a programme has no timetable templates
        # at all. This avoids one-off records showing as fake weekly classes.
        if not rows and not programme_has_templates:
            weekday_sql = day_date.strftime("%w")
            recent_cutoff = (today - timedelta(days=210)).isoformat()
            derived_rows = db.execute("""
                SELECT
                    NULLIF(trim(class_name), '') AS class_name_raw,
                    COALESCE(NULLIF(trim(class_name), ''), 'General Class') AS class_name_display,
                    COALESCE(NULLIF(trim(session), ''), '') AS session,
                    COALESCE(NULLIF(trim(location), ''), ?) AS location,
                    COUNT(*) AS seen_count
                FROM tasters
                WHERE programme=?
                  AND strftime('%w', taster_date)=?
                  AND taster_date>=?
                GROUP BY class_name, session, location
                ORDER BY seen_count DESC, session, class_name
            """, (
                programme.title(),
                programme,
                weekday_sql,
                recent_cutoff,
            )).fetchall()
            projected = []
            seen_keys = set()
            for drow in derived_rows:
                session_raw = (drow["session"] or "").strip()
                start_time = normalise_session_label(session_raw)
                if ":" not in start_time:
                    start_time = _extract_time(session_raw)
                if not start_time:
                    continue
                key = (drow["class_name_display"], start_time)
                if key in seen_keys:
                    continue
                seen_keys.add(key)
                projected.append({
                    "class_name": drow["class_name_display"],
                    "match_class_name": drow["class_name_raw"],
                    "start_time": start_time,
                    "end_time": "",
                    "location": drow["location"] or programme.title(),
                })
            if projected:
                rows = projected
                source_mode = "derived"

        sessions = []
        for row in rows:
            start_time = str(row["start_time"] or "")[:5]
            end_time = str(row["end_time"] or "")[:5]
            time_range = f"{start_time} - {end_time}" if end_time else start_time
            weekday_sql = day_date.strftime("%w")
            session_time = normalise_session_label(start_time)
            session_with_day = f"{day_name} {start_time}".strip()
            class_name = str(row["class_name"] or "").strip() or "General Class"
            class_name_match = (
                str(row["match_class_name"]).strip()
                if isinstance(row, dict) and row.get("match_class_name") is not None
                else class_name
            )
            if class_name_match.lower() == "general class":
                class_name_match = ""

            if class_name_match:
                if session_time:
                    upcoming_count = db.execute("""
                        SELECT COUNT(*) AS c
                        FROM tasters
                        WHERE programme=?
                          AND class_name=?
                          AND (
                            lower(trim(session)) = lower(?)
                            OR lower(trim(session)) = lower(?)
                          )
                          AND strftime('%w', taster_date)=?
                          AND taster_date>=?
                          AND taster_date<=?
                    """, (
                        programme,
                        class_name_match,
                        session_time,
                        session_with_day,
                        weekday_sql,
                        window_start,
                        window_end
                    )).fetchone()["c"]
                else:
                    upcoming_count = db.execute("""
                        SELECT COUNT(*) AS c
                        FROM tasters
                        WHERE programme=?
                          AND class_name=?
                          AND strftime('%w', taster_date)=?
                          AND taster_date>=?
                          AND taster_date<=?
                    """, (
                        programme,
                        class_name_match,
                        weekday_sql,
                        window_start,
                        window_end
                    )).fetchone()["c"]
            else:
                if session_time:
                    upcoming_count = db.execute("""
                        SELECT COUNT(*) AS c
                        FROM tasters
                        WHERE programme=?
                          AND (
                            lower(trim(session)) = lower(?)
                            OR lower(trim(session)) = lower(?)
                          )
                          AND strftime('%w', taster_date)=?
                          AND taster_date>=?
                          AND taster_date<=?
                    """, (
                        programme,
                        session_time,
                        session_with_day,
                        weekday_sql,
                        window_start,
                        window_end
                    )).fetchone()["c"]
                else:
                    upcoming_count = db.execute("""
                        SELECT COUNT(*) AS c
                        FROM tasters
                        WHERE programme=?
                          AND strftime('%w', taster_date)=?
                          AND taster_date>=?
                          AND taster_date<=?
                    """, (
                        programme,
                        weekday_sql,
                        window_start,
                        window_end
                    )).fetchone()["c"]

            sessions.append({
                "session_value": f"{day_name} {start_time}",
                "class_name": class_name,
                "class_type": infer_class_type(class_name, start_time, end_time),
                "time_range": time_range,
                "start_time": start_time,
                "end_time": end_time,
                "location": row["location"],
                "upcoming_count": int(upcoming_count or 0),
            })

        week_days.append({
            "date_obj": day_date,
            "date_str": day_str,
            "day_name": day_name,
            "sessions": sessions,
            "source_mode": source_mode,
        })

    return week_days


def toggle_flag(taster_id, column):
    if column not in ("attended", "club_fees", "bg", "badge"):
        return None
    db = get_db()
    cur = db.cursor()
    cur.execute(f"SELECT {column} FROM tasters WHERE id=?", (taster_id,))
    row = cur.fetchone()
    if not row:
        return None
    new_value = 0 if row[column] else 1
    cur.execute(
        f"UPDATE tasters SET {column}=? WHERE id=?",
        (new_value, taster_id)
    )
    db.commit()
    updated = cur.execute("SELECT * FROM tasters WHERE id=?", (taster_id,)).fetchone()
    return dict(updated) if updated else None

# ==========================================================
# DASHBOARD (HOME)
# ==========================================================

@app.route("/")
@app.route("/dashboard")
def dashboard():
    db = get_db()
    today = date.today()
    month_key = today.strftime("%Y-%m")
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    cutoff_iso = (today - timedelta(days=62)).isoformat()
    days_in_month = calendar.monthrange(today.year, today.month)[1]
    days_left = max(days_in_month - today.day, 0)

    tasters_month = db.execute("""
        SELECT COUNT(*) c FROM tasters
        WHERE strftime('%Y-%m', taster_date)=?
    """, (month_key,)).fetchone()["c"]

    leavers_month = db.execute("""
        SELECT COUNT(*) c FROM leavers
        WHERE leave_month=?
    """, (month_key,)).fetchone()["c"]
    total_tasters_all = db.execute(
        "SELECT COUNT(*) c FROM tasters"
    ).fetchone()["c"]

    todays_counts_rows = db.execute("""
        SELECT programme, COUNT(*) c
        FROM tasters
        WHERE taster_date=?
        GROUP BY programme
    """, (today.isoformat(),)).fetchall()
    todays_counts = {"lockwood": 0, "honley": 0, "preschool": 0}
    for row in todays_counts_rows:
        todays_counts[row["programme"]] = int(row["c"] or 0)
    todays_total = sum(todays_counts.values())

    todays_tasters_rows = db.execute("""
        SELECT child, programme, session, class_name
        FROM tasters
        WHERE taster_date=?
        ORDER BY session, child
    """, (today.isoformat(),)).fetchall()
    todays_by_programme = {
        "lockwood": [],
        "honley": [],
        "preschool": [],
    }
    for row in todays_tasters_rows:
        programme = row["programme"] if row["programme"] in todays_by_programme else "lockwood"
        todays_by_programme[programme].append(dict(row))

    month_by_programme_rows = db.execute("""
        SELECT programme, COUNT(*) c
        FROM tasters
        WHERE strftime('%Y-%m', taster_date)=?
        GROUP BY programme
    """, (month_key,)).fetchall()
    counts = {"lockwood": 0, "honley": 0, "preschool": 0}
    for row in month_by_programme_rows:
        counts[row["programme"]] = row["c"]
    max_count = max(counts.values()) if counts else 0
    month_by_programme = []
    for key, label in (
        ("lockwood", "Lockwood"),
        ("honley", "Honley"),
        ("preschool", "Preschool"),
    ):
        count = counts.get(key, 0)
        pct = int((count / max_count) * 100) if max_count > 0 else 0
        month_by_programme.append({
            "key": key,
            "label": label,
            "count": count,
            "pct": pct,
        })

    followups_open = db.execute("""
        SELECT COUNT(*) c
        FROM tasters
        WHERE taster_date>=?
          AND taster_date<=?
          AND (attended=0 OR club_fees=0 OR bg=0 OR badge=0)
    """, (cutoff_iso, today.isoformat())).fetchone()["c"]

    converted_month = db.execute("""
        SELECT COUNT(*) c
        FROM tasters
        WHERE strftime('%Y-%m', taster_date)=?
          AND attended=1
          AND club_fees=1
          AND bg=1
          AND badge=1
    """, (month_key,)).fetchone()["c"]

    week_bookings = db.execute("""
        SELECT COUNT(*) c
        FROM tasters
        WHERE taster_date>=?
          AND taster_date<=?
    """, (week_start.isoformat(), week_end.isoformat())).fetchone()["c"]

    last_import = load_last_import_data()
    monitor = {
        "status": "idle",
        "label": "Not Run",
        "class_name": "monitor-idle",
        "run_at": "—",
        "warnings": 0,
        "import_total_tasters": None,
        "db_total_tasters": total_tasters_all,
    }
    if last_import:
        monitor["run_at"] = last_import.get("run_at", "—")
        monitor["warnings"] = len(last_import.get("warnings", []))
        monitor["import_total_tasters"] = last_import.get("total_tasters")
        if last_import.get("exit_code") != 0:
            monitor["status"] = "error"
            monitor["label"] = "Import Failed"
            monitor["class_name"] = "monitor-error"
        elif monitor["warnings"] > 0:
            monitor["status"] = "warn"
            monitor["label"] = "Check Import"
            monitor["class_name"] = "monitor-warn"
        else:
            monitor["status"] = "ok"
            monitor["label"] = "Import Healthy"
            monitor["class_name"] = "monitor-ok"

    return render_template(
        "dashboard.html",
        month=today.strftime("%B %Y"),
        current_time=datetime.now().strftime("%H:%M"),
        today=today,
        tasters=tasters_month,
        leavers=leavers_month,
        net=tasters_month - leavers_month,
        followups_open=followups_open,
        converted_month=converted_month,
        week_bookings=week_bookings,
        todays_total=todays_total,
        todays_counts=todays_counts,
        todays_by_programme=todays_by_programme,
        month_by_programme=month_by_programme,
        monitor=monitor,
        days_left=days_left,
        week_start=week_start,
        week_end=week_end,
    )


# ==========================================================
# CORE VIEWS
# ==========================================================

@app.route("/today")
def today():
    programme = request.args.get("programme", "lockwood")
    return redirect(url_for(
        "day_detail",
        date_str=date.today().isoformat(),
        programme=programme
    ))

@app.post("/toggle/<int:taster_id>/<field>", endpoint="toggle")
def toggle_view(taster_id, field):
    if field not in ("attended", "club_fees", "bg", "badge"):
        flash("Invalid toggle field", "danger")
        return redirect(request.referrer or url_for("today"))

    updated_row = toggle_flag(taster_id, field)
    if updated_row:
        initials = user_initials((current_user() or {}).get("full_name", ""))
        ok, msg = sync_taster_to_excel(updated_row, mode="status", changed_field=field, actor_initials=initials)
        log_audit(
            "toggle_taster_field",
            entity_type="taster",
            entity_id=taster_id,
            details=f"{field}={updated_row.get(field)} | excel_sync={msg}",
            status="ok" if ok else "warn",
        )
        if not ok:
            flash(f"Updated in app, but Excel sync needs review: {msg}", "warning")
    else:
        log_audit(
            "toggle_taster_field",
            entity_type="taster",
            entity_id=taster_id,
            details=f"Toggle failed: taster not found ({field})",
            status="warn",
        )
    return redirect(request.referrer or url_for("today"))

@app.route("/month")
def month():
    programme = request.args.get("programme", "lockwood")
    year = int(request.args.get("y", date.today().year))
    month_n = int(request.args.get("m", date.today().month))

    df = load_tasters_df(programme)
    if not df.empty:
        df = df[
            (pd.to_datetime(df["taster_date"]).dt.year == year) &
            (pd.to_datetime(df["taster_date"]).dt.month == month_n)
        ]

    db = get_db()
    year_rows = db.execute("""
        SELECT DISTINCT CAST(strftime('%Y', taster_date) AS INTEGER) AS y
        FROM tasters
        WHERE taster_date IS NOT NULL
        ORDER BY y
    """).fetchall()
    year_options = [row["y"] for row in year_rows if row["y"]]
    if not year_options:
        current = date.today().year
        year_options = list(range(current - 1, current + 2))
    if year not in year_options:
        year_options.append(year)
        year_options = sorted(set(year_options))

    return render_template(
        "month.html",
        year=year,
        month=month_n,
        month_name=calendar.month_name[month_n],
        month_matrix=calendar.monthcalendar(year, month_n),
        month_df=df,
        programme=programme,
        location=programme.title(),
        today_date=date.today(),
        year_options=year_options,
        calendar=calendar,
        datetime=datetime
    )


@app.route("/day/<date_str>")
def day_detail(date_str):
    programme = request.args.get("programme", "lockwood")
    try:
        selected = datetime.fromisoformat(date_str).date()
    except ValueError:
        flash("Invalid date", "danger")
        return redirect(url_for("dashboard"))

    df = load_tasters_df(programme)
    day_df = df[df["taster_date"] == selected] if not df.empty else df
    day_stats = {
        "total": 0,
        "attended": 0,
        "club_fees": 0,
        "bg": 0,
        "badge": 0,
        "fully_complete": 0,
        "to_action": 0,
        "completion_pct": 0,
    }
    session_totals = []

    if not day_df.empty:
        attended_flag = pd.to_numeric(day_df["attended"], errors="coerce").fillna(0).astype(int).clip(0, 1)
        fees_flag = pd.to_numeric(day_df["club_fees"], errors="coerce").fillna(0).astype(int).clip(0, 1)
        bg_flag = pd.to_numeric(day_df["bg"], errors="coerce").fillna(0).astype(int).clip(0, 1)
        badge_flag = pd.to_numeric(day_df["badge"], errors="coerce").fillna(0).astype(int).clip(0, 1)
        complete_mask = (attended_flag == 1) & (fees_flag == 1) & (bg_flag == 1) & (badge_flag == 1)

        total_rows = int(len(day_df.index))
        fully_complete = int(complete_mask.sum())
        to_action = max(0, total_rows - fully_complete)
        completion_pct = int(round((fully_complete / total_rows) * 100)) if total_rows else 0

        day_stats = {
            "total": total_rows,
            "attended": int(attended_flag.sum()),
            "club_fees": int(fees_flag.sum()),
            "bg": int(bg_flag.sum()),
            "badge": int(badge_flag.sum()),
            "fully_complete": fully_complete,
            "to_action": to_action,
            "completion_pct": completion_pct,
        }

        grouped_df = day_df.copy()
        grouped_df["_attended"] = attended_flag
        grouped_df["_fees"] = fees_flag
        grouped_df["_bg"] = bg_flag
        grouped_df["_badge"] = badge_flag

        for session_value, group in grouped_df.groupby("session", dropna=False):
            label = ""
            if pd.notna(session_value):
                label = str(session_value).strip()
            if not label:
                label = selected.strftime("%A")

            group_complete = int(
                (
                    (group["_attended"] == 1)
                    & (group["_fees"] == 1)
                    & (group["_bg"] == 1)
                    & (group["_badge"] == 1)
                ).sum()
            )
            group_total = int(len(group.index))
            group_to_action = max(0, group_total - group_complete)
            group_pct = int(round((group_complete / group_total) * 100)) if group_total else 0
            session_totals.append(
                {
                    "label": label,
                    "total": group_total,
                    "complete": group_complete,
                    "to_action": group_to_action,
                    "pct": group_pct,
                }
            )

    return render_template(
        "day.html",
        data=day_df,
        day_stats=day_stats,
        session_totals=session_totals,
        selected_date=selected,
        prev_date=(selected - timedelta(days=1)),
        next_date=(selected + timedelta(days=1)),
        today_date=date.today(),
        programme=programme,
        location=programme.title()
    )

@app.route("/stats")
def stats():
    raw_monthly = query("""
        WITH t AS (
            SELECT strftime('%Y-%m', taster_date) AS m, COUNT(*) c
            FROM tasters
            GROUP BY m
        ),
        l AS (
            SELECT leave_month AS m, COUNT(*) c
            FROM leavers
            GROUP BY m
        )
        SELECT
            COALESCE(t.m, l.m) AS month,
            COALESCE(t.c, 0) AS tasters,
            COALESCE(l.c, 0) AS leavers
        FROM t
        LEFT JOIN l ON t.m = l.m
        UNION
        SELECT
            COALESCE(t.m, l.m),
            COALESCE(t.c, 0),
            COALESCE(l.c, 0)
        FROM l
        LEFT JOIN t ON t.m = l.m
        ORDER BY month;
    """)
    monthly = []
    for row in raw_monthly:
        m = (row["month"] or "").strip()
        if not re.match(r"^\d{4}-\d{2}$", m):
            continue
        if m < "2000-01":
            continue
        monthly.append({
            "month": m,
            "tasters": int(row["tasters"] or 0),
            "leavers": int(row["leavers"] or 0),
        })
    monthly_desc = sorted(monthly, key=lambda r: r["month"], reverse=True)
    monthly_chart = sorted(monthly, key=lambda r: r["month"])
    current_month = date.today().strftime("%Y-%m")
    monthly_desc = [r for r in monthly_desc if r["month"] <= current_month]
    monthly_chart = [r for r in monthly_chart if r["month"] <= current_month]

    this_month = date.today().strftime("%Y-%m")
    month_programme_rows = query("""
        SELECT programme, COUNT(*) AS c
        FROM tasters
        WHERE strftime('%Y-%m', taster_date)=?
        GROUP BY programme
        ORDER BY c DESC
    """, (this_month,))
    this_month_programme = [
        {"programme": r["programme"], "count": int(r["c"] or 0)}
        for r in month_programme_rows
    ]

    totals = {
        "tasters_all": sum(r["tasters"] for r in monthly_desc),
        "leavers_all": sum(r["leavers"] for r in monthly_desc),
        "months_tracked": len(monthly_desc),
        "net_all": sum(r["tasters"] - r["leavers"] for r in monthly_desc),
        "latest_month": monthly_desc[0]["month"] if monthly_desc else "—",
    }

    return render_template(
        "stats.html",
        monthly=monthly_desc,
        monthly_chart=monthly_chart,
        current_month=current_month,
        totals=totals,
        this_month_programme=this_month_programme,
    )


# ==========================================================
# LEAVERS
# ==========================================================

@app.route("/leavers/add", methods=["GET", "POST"])
def add_leaver():
    programme = request.args.get("programme", "preschool")
    if request.method == "POST":
        child = normalise_child_name(request.form["child"])
        programme = request.form.get("programme", programme)
        leave_date = request.form.get("leave_date", "").strip()
        session_label = request.form.get("session", "").strip()
        class_day = request.form.get("class_day", "").strip()
        class_name = request.form.get("class_name", "").strip()
        removed_la = 1 if request.form.get("removed_la") == "1" else 0
        removed_bg = 1 if request.form.get("removed_bg") == "1" else 0
        added_to_board = 1 if request.form.get("added_to_board") == "1" else 0
        reason = request.form.get("reason", "").strip()
        email = request.form.get("email", "").strip()
        sync_excel = False

        if not child or not leave_date:
            flash("Name and leave date are required.", "danger")
            return redirect(request.url)

        if not class_day:
            class_day = extract_day_name(session_label)
        session_label = normalise_session_label(session_label)

        if not session_label:
            flash("Please choose a class/session.", "warning")
            return redirect(request.url)

        try:
            leave_dt = datetime.fromisoformat(leave_date).date()
            leave_month = leave_dt.strftime("%Y-%m")
            if not class_day:
                class_day = leave_dt.strftime("%A")
        except ValueError:
            flash("Invalid leave date", "danger")
            return redirect(request.url)

        db = get_db()
        db.execute("""
            INSERT INTO leavers (
                child, programme, leave_month, leave_date,
                class_day, session, class_name,
                removed_la, removed_bg, added_to_board, reason,
                email, source
            )
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            child,
            programme,
            leave_month,
            leave_date,
            class_day,
            session_label,
            class_name,
            removed_la,
            removed_bg,
            added_to_board,
            reason,
            email,
            "manual"
        ))
        leaver_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        db.commit()

        inserted = db.execute("SELECT * FROM leavers WHERE id=?", (leaver_id,)).fetchone()
        actor_initials = user_initials((current_user() or {}).get("full_name", ""))
        sync_msg = "Excel sync skipped"
        sync_status = "ok"
        if sync_excel and inserted:
            ok, sync_msg = sync_leaver_to_excel(inserted, actor_initials=actor_initials)
            if not ok:
                flash(f"Leaver saved in app, but Excel sync needs review: {sync_msg}", "warning")
                sync_status = "warn"

        log_audit(
            "add_leaver",
            entity_type="leaver",
            entity_id=leaver_id,
            details=f"{child} | {programme} | {class_day} {session_label} | {leave_date} | excel_sync={sync_msg}",
            status=sync_status,
        )

        flash(f"Leaver recorded for {child}", "success")
        return redirect(url_for("admin_tasks"))

    week_start_raw = request.args.get("week_start") or request.args.get("leave_date")
    if week_start_raw:
        try:
            anchor_date = datetime.fromisoformat(week_start_raw).date()
        except ValueError:
            anchor_date = date.today()
    else:
        anchor_date = date.today()

    week_start = anchor_date - timedelta(days=anchor_date.weekday())
    week_end = week_start + timedelta(days=6)
    week_days = build_week_schedule(programme, week_start)

    return render_template(
        "add_leaver.html",
        programme=programme,
        week_days=week_days,
        week_start=week_start,
        week_end=week_end,
        prev_week=(week_start - timedelta(days=7)).isoformat(),
        next_week=(week_start + timedelta(days=7)).isoformat(),
        today_str=date.today().isoformat()
    )


@app.route("/leavers/add/manual", methods=["GET", "POST"])
def add_manual_leaver():
    programme = request.args.get("programme", "preschool")

    if request.method == "POST":
        child = normalise_child_name(request.form["child"])
        programme = request.form.get("programme", programme).strip().lower() or programme
        leave_date = request.form.get("leave_date", "").strip()
        class_name = request.form.get("class_name", "").strip() or "Manual Session"
        class_day = request.form.get("class_day", "").strip()
        session_label = normalise_session_label(request.form.get("session_label", "").strip())
        removed_la = 1 if request.form.get("removed_la") == "1" else 0
        removed_bg = 1 if request.form.get("removed_bg") == "1" else 0
        added_to_board = 1 if request.form.get("added_to_board") == "1" else 0
        reason = request.form.get("reason", "").strip()
        email = request.form.get("email", "").strip()
        sync_excel = False

        if not child or not leave_date or not session_label:
            flash("Name, leave date, and session label are required.", "danger")
            return redirect(request.url)

        try:
            leave_dt = datetime.fromisoformat(leave_date).date()
            leave_month = leave_dt.strftime("%Y-%m")
        except ValueError:
            flash("Invalid leave date", "danger")
            return redirect(request.url)

        if not class_day:
            class_day = extract_day_name(session_label) or leave_dt.strftime("%A")

        db = get_db()
        db.execute("""
            INSERT INTO leavers (
                child, programme, leave_month, leave_date,
                class_day, session, class_name,
                removed_la, removed_bg, added_to_board, reason,
                email, source
            )
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            child,
            programme,
            leave_month,
            leave_date,
            class_day,
            session_label,
            class_name,
            removed_la,
            removed_bg,
            added_to_board,
            reason,
            email,
            "manual"
        ))
        leaver_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        db.commit()

        inserted = db.execute("SELECT * FROM leavers WHERE id=?", (leaver_id,)).fetchone()
        actor_initials = user_initials((current_user() or {}).get("full_name", ""))
        sync_msg = "Excel sync skipped"
        sync_status = "ok"
        if sync_excel and inserted:
            ok, sync_msg = sync_leaver_to_excel(inserted, actor_initials=actor_initials)
            if not ok:
                flash(f"Leaver saved in app, but Excel sync needs review: {sync_msg}", "warning")
                sync_status = "warn"

        log_audit(
            "add_leaver_manual",
            entity_type="leaver",
            entity_id=leaver_id,
            details=f"{child} | {programme} | {class_day} {session_label} | {leave_date} | excel_sync={sync_msg}",
            status=sync_status,
        )
        flash(f"Manual leaver recorded for {child}", "success")
        return redirect(url_for("admin_tasks"))

    week_start_raw = request.args.get("week_start") or request.args.get("leave_date")
    if week_start_raw:
        try:
            anchor_date = datetime.fromisoformat(week_start_raw).date()
        except ValueError:
            anchor_date = date.today()
    else:
        anchor_date = date.today()
    week_start = anchor_date - timedelta(days=anchor_date.weekday())
    today_str = date.today().isoformat()
    return render_template(
        "add_leaver_manual.html",
        programme=programme,
        week_start=week_start,
        today_str=today_str,
    )


@app.route("/admin/tasks")
def admin_tasks():
    today_dt = date.today()
    today_iso = today_dt.isoformat()
    cutoff_iso = three_month_cutoff_date(today_dt).isoformat()
    month_key = today_dt.strftime("%Y-%m")
    month_label = today_dt.strftime("%B %Y")
    user = current_user()
    assignments = user.get("admin_days", []) if user else []
    assignments_sorted = sorted(
        assignments,
        key=lambda a: (DAY_ORDER.get(a["day_name"], 99), a["programme"])
    )
    assignment_set = {(a["day_name"], a["programme"]) for a in assignments_sorted}

    def key_for(day_name, programme):
        return f"{day_name}|{programme}"

    db = get_db()
    leaver_count_map = {}
    unknown_leaver_counts = {}
    leaver_rows = db.execute("""
        SELECT child, programme, leave_date, class_day, session, class_name
        FROM leavers
        WHERE leave_month=?
    """, (month_key,)).fetchall()
    for row in leaver_rows:
        row_dict = dict(row)
        day_name = (
            extract_day_name(row_dict.get("class_day"))
            or extract_day_name(row_dict.get("session"))
        )
        leave_date = str(row_dict.get("leave_date") or "").strip()
        if not day_name and leave_date:
            try:
                day_name = datetime.fromisoformat(leave_date).strftime("%A")
            except ValueError:
                day_name = ""
        if not day_name:
            unknown_key = key_for("?", row_dict["programme"])
            unknown_leaver_counts[unknown_key] = unknown_leaver_counts.get(unknown_key, 0) + 1
            continue
        if assignment_set and (day_name, row_dict["programme"]) not in assignment_set:
            continue
        key = key_for(day_name, row_dict["programme"])
        leaver_count_map[key] = leaver_count_map.get(key, 0) + 1

    member_count_map = {}
    member_rows = db.execute("""
        SELECT taster_date, programme
        FROM tasters
        WHERE strftime('%Y-%m', taster_date)=?
          AND attended=1
          AND club_fees=1
          AND bg=1
          AND badge=1
    """, (month_key,)).fetchall()
    for row in member_rows:
        day_name = datetime.fromisoformat(row["taster_date"]).strftime("%A")
        if assignment_set and (day_name, row["programme"]) not in assignment_set:
            continue
        key = key_for(day_name, row["programme"])
        member_count_map[key] = member_count_map.get(key, 0) + 1

    followup_rows = []
    raw_followups = db.execute("""
        SELECT
            id, child, programme, taster_date, session, class_name,
            attended, club_fees, bg, badge, reschedule_contacted, notes
        FROM tasters
        WHERE taster_date>=?
          AND taster_date<=?
          AND (attended=0 OR club_fees=0 OR bg=0 OR badge=0)
        ORDER BY taster_date DESC, programme, session, child
    """, (cutoff_iso, today_iso)).fetchall()
    for row in raw_followups:
        row_dict = dict(row)
        day_name = datetime.fromisoformat(row["taster_date"]).strftime("%A")
        if assignment_set and (day_name, row["programme"]) not in assignment_set:
            continue
        row_dict["day_name"] = day_name
        followup_rows.append(row_dict)

    if assignments_sorted:
        summary_rows = []
        for a in assignments_sorted:
            key = key_for(a["day_name"], a["programme"])
            summary_rows.append({
                "day_name": a["day_name"],
                "programme": a["programme"],
                "key": key,
                "leavers": leaver_count_map.get(key, 0),
                "members": member_count_map.get(key, 0),
            })
    else:
        keys = set(leaver_count_map.keys()) | set(member_count_map.keys())
        for row in followup_rows:
            keys.add(key_for(row["day_name"], row["programme"]))
        summary_rows = []
        for key in sorted(
            keys,
            key=lambda k: (
                DAY_ORDER.get(k.split("|", 1)[0], 99),
                k.split("|", 1)[1]
            )
        ):
            day_name, programme = key.split("|", 1)
            summary_rows.append({
                "day_name": day_name,
                "programme": programme,
                "key": key,
                "leavers": leaver_count_map.get(key, 0),
                "members": member_count_map.get(key, 0),
            })

    unknown_summary = []
    for key, count in sorted(unknown_leaver_counts.items()):
        _, programme = key.split("|", 1)
        unknown_summary.append({
            "programme": programme,
            "count": count,
        })

    return render_template(
        "admin_tasks.html",
        month_label=month_label,
        summary_rows=summary_rows,
        unknown_summary=unknown_summary,
        followup_rows=followup_rows,
        followup_total=len(followup_rows),
        assignments=assignments_sorted,
        has_custom_assignments=bool(assignment_set),
        today=today_iso,
        cutoff=cutoff_iso,
    )


@app.post("/admin/tasks/contact/<int:taster_id>")
def admin_mark_contacted(taster_id):
    db = get_db()
    row = db.execute("SELECT * FROM tasters WHERE id=?", (taster_id,)).fetchone()
    if not row:
        flash("Taster not found.", "warning")
        return redirect(request.referrer or url_for("admin_tasks"))
    db.execute("UPDATE tasters SET reschedule_contacted=1 WHERE id=?", (taster_id,))
    db.commit()
    updated_row = db.execute("SELECT * FROM tasters WHERE id=?", (taster_id,)).fetchone()
    sync_ok = True
    sync_msg = "Excel sync skipped"
    if updated_row:
        initials = user_initials((current_user() or {}).get("full_name", ""))
        sync_ok, sync_msg = sync_taster_to_excel(
            updated_row,
            mode="contacted",
            actor_initials=initials
        )
        if not sync_ok:
            flash(f"Marked contacted in app, but Excel sync needs review: {sync_msg}", "warning")
    log_audit(
        "mark_no_show_contacted",
        entity_type="taster",
        entity_id=taster_id,
        details=f"Marked as contacted for reschedule | excel_sync={sync_msg}",
        status="ok" if sync_ok else "warn",
    )
    flash("Marked as contacted.", "success")
    return redirect(request.referrer or url_for("admin_tasks"))


@app.route("/account", methods=["GET", "POST"])
def account_settings():
    user = current_user()
    if user is None:
        return redirect(url_for("login"))

    db = get_db()
    if request.method == "POST":
        action = request.form.get("action", "")
        if action == "profile":
            first_name = request.form.get("first_name", "").strip()
            last_name = request.form.get("last_name", "").strip()
            email = request.form.get("email", "").strip().lower()
            full_name = f"{first_name} {last_name}".strip()
            if not first_name or not last_name:
                flash("First and last name are required.", "warning")
                return redirect(url_for("account_settings"))
            if not email or "@" not in email:
                flash("Valid email is required.", "warning")
                return redirect(url_for("account_settings"))
            existing = db.execute(
                "SELECT id FROM users WHERE username=? AND id<>?",
                (email, user["id"])
            ).fetchone()
            if existing:
                flash("That email is already in use.", "warning")
                return redirect(url_for("account_settings"))
            db.execute(
                "UPDATE users SET full_name=?, username=? WHERE id=?",
                (full_name, email, user["id"])
            )
            db.commit()
            log_audit(
                "update_profile",
                entity_type="user",
                entity_id=user["id"],
                details=f"Profile updated to {full_name} ({email})",
            )
            flash("Profile updated.", "success")
            return redirect(url_for("account_settings"))
        if action == "password":
            current_password = request.form.get("current_password", "")
            new_password = request.form.get("new_password", "")
            confirm_password = request.form.get("confirm_password", "")
            row = db.execute(
                "SELECT password_hash FROM users WHERE id=?",
                (user["id"],)
            ).fetchone()
            require_current_password = not is_admin_user(user)
            if require_current_password and (not row or not check_password_hash(row["password_hash"], current_password)):
                flash("Current password is incorrect.", "danger")
                return redirect(url_for("account_settings"))
            if enforce_password_policy() and not is_admin_user(user):
                password_errors = password_strength_errors(new_password)
                if password_errors:
                    flash("New password " + "; ".join(password_errors) + ".", "warning")
                    return redirect(url_for("account_settings"))
            if new_password != confirm_password:
                flash("New password and confirmation do not match.", "warning")
                return redirect(url_for("account_settings"))
            db.execute(
                "UPDATE users SET password_hash=?, password_must_change=0 WHERE id=?",
                (generate_password_hash(new_password), user["id"])
            )
            db.commit()
            session.pop("must_change_password", None)
            log_audit("change_password", entity_type="user", entity_id=user["id"], details="Password updated")
            flash("Password updated.", "success")
            return redirect(url_for("account_settings"))

        if action == "email_prefs":
            weekly_opt_in = 1 if request.form.get("weekly_report_opt_in") == "1" else 0
            db.execute(
                "UPDATE users SET email_weekly_reports=? WHERE id=?",
                (weekly_opt_in, user["id"]),
            )
            db.commit()
            log_audit(
                "update_email_prefs",
                entity_type="user",
                entity_id=user["id"],
                details=f"weekly_report_opt_in={weekly_opt_in}",
            )
            if email_owner_only_mode() and user.get("username", "").lower() != OWNER_EMAIL:
                flash("Saved. Owner-only mode is active, so weekly emails still only go to owner.", "warning")
            else:
                flash("Email preferences updated.", "success")
            return redirect(url_for("account_settings"))

        if action == "admin_days":
            selected_values = parse_admin_day_values(request.form.getlist("admin_days"))
            db.execute("DELETE FROM user_admin_days WHERE user_id=?", (user["id"],))
            for day_name, programme in selected_values:
                db.execute("""
                    INSERT INTO user_admin_days (user_id, day_name, programme)
                    VALUES (?, ?, ?)
                    ON CONFLICT(user_id, day_name, programme) DO NOTHING
                """, (user["id"], day_name, programme))
            db.commit()
            log_audit(
                "update_admin_days",
                entity_type="user",
                entity_id=user["id"],
                details=f"Selected {len(selected_values)} admin day cells",
            )
            flash("Admin day ownership updated.", "success")
            return redirect(url_for("account_settings"))

    selected_set = {
        f"{r['day_name']}|{r['programme']}"
        for r in db.execute(
            "SELECT day_name, programme FROM user_admin_days WHERE user_id=?",
            (user["id"],)
        ).fetchall()
        if admin_day_cell_allowed(r["day_name"], r["programme"])
    }

    grouped_options = build_admin_day_grouped_options()

    full_name_parts = (user.get("full_name") or "").strip().split()
    first_name = full_name_parts[0] if full_name_parts else ""
    last_name = " ".join(full_name_parts[1:]) if len(full_name_parts) > 1 else ""

    return render_template(
        "account.html",
        user=user,
        first_name=first_name,
        last_name=last_name,
        weekly_report_opt_in=bool(user.get("email_weekly_reports")),
        selected_set=selected_set,
        grouped_options=grouped_options,
        last_import=load_last_import_data()
    )


@app.route("/account/admin", methods=["GET", "POST"])
@owner_required
def account_admin():
    db = get_db()
    admin_user = current_user()

    if request.method == "POST":
        action = request.form.get("action", "").strip()
        if action == "create_user":
            full_name = request.form.get("full_name", "").strip()
            username = request.form.get("username", "").strip().lower()
            role = request.form.get("role", "staff").strip().lower()
            selected_values = parse_admin_day_values(request.form.getlist("admin_days"))

            if not full_name or not username:
                flash("Name and email are required.", "warning")
                return redirect(url_for("account_admin"))
            if role not in {"admin", "staff"}:
                role = "staff"

            existing = db.execute(
                "SELECT id FROM users WHERE username=?",
                (username,)
            ).fetchone()
            if existing:
                flash("That email is already used by another account.", "warning")
                return redirect(url_for("account_admin"))

            default_password = os.environ.get("TASTERIST_DEFAULT_USER_PASSWORD", "JamesRocks1946!").strip()
            temporary_password = default_password or secrets.token_urlsafe(10)
            db.execute("""
                INSERT INTO users (username, password_hash, full_name, role, password_must_change)
                VALUES (?, ?, ?, ?, 0)
            """, (username, generate_password_hash(temporary_password), full_name, role))
            target_user_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]

            for day_name, programme in selected_values:
                db.execute("""
                    INSERT INTO user_admin_days (user_id, day_name, programme)
                    VALUES (?, ?, ?)
                    ON CONFLICT(user_id, day_name, programme) DO NOTHING
                """, (target_user_id, day_name, programme))
            db.commit()
            log_audit(
                "admin_create_user",
                entity_type="user",
                entity_id=target_user_id,
                details=f"Created account {username} role={role} admin_days={len(selected_values)}",
            )
            flash(
                f"Created account: {username}. Password is {temporary_password}",
                "success",
            )
            return redirect(url_for("account_admin"))

        user_id_raw = request.form.get("user_id", "").strip()
        try:
            target_user_id = int(user_id_raw)
        except ValueError:
            flash("Invalid account selection.", "warning")
            return redirect(url_for("account_admin"))

        target_user = db.execute(
            "SELECT id, username, full_name, role FROM users WHERE id=?",
            (target_user_id,)
        ).fetchone()
        if not target_user:
            flash("Account not found.", "warning")
            return redirect(url_for("account_admin"))

        if action == "delete_user":
            if admin_user and target_user_id == admin_user["id"]:
                flash("You cannot remove your own account.", "warning")
                return redirect(url_for("account_admin"))

            target_role = (target_user["role"] or "").lower()
            if target_role == "owner":
                flash("Owner account cannot be removed.", "warning")
                return redirect(url_for("account_admin"))

            if target_role == "admin":
                admin_count = db.execute(
                    "SELECT COUNT(*) AS c FROM users WHERE lower(role) IN ('admin','owner')"
                ).fetchone()["c"]
                if admin_count <= 1:
                    flash("Cannot remove the last admin-level account.", "warning")
                    return redirect(url_for("account_admin"))

            db.execute("DELETE FROM user_admin_days WHERE user_id=?", (target_user_id,))
            db.execute("DELETE FROM users WHERE id=?", (target_user_id,))
            db.commit()
            log_audit(
                "admin_delete_user",
                entity_type="user",
                entity_id=target_user_id,
                details=f"Deleted account {target_user['username']}",
                status="warn",
            )
            flash(f"Removed account: {target_user['username']}", "success")
            return redirect(url_for("account_admin"))

        if action == "save_user":
            full_name = request.form.get("full_name", "").strip()
            username = request.form.get("username", "").strip().lower()
            role = request.form.get("role", "staff").strip().lower()
            new_password = request.form.get("new_password", "")
            selected_values = parse_admin_day_values(request.form.getlist("admin_days"))

            if not full_name or not username:
                flash("Name and email are required.", "warning")
                return redirect(url_for("account_admin"))

            if role not in {"admin", "staff"}:
                role = "staff"

            target_role = (target_user["role"] or "").lower()
            if target_role == "owner":
                role = "owner"

            username_conflict = db.execute(
                "SELECT id FROM users WHERE username=? AND id<>?",
                (username, target_user_id)
            ).fetchone()
            if username_conflict:
                flash("That email is already used by another account.", "warning")
                return redirect(url_for("account_admin"))

            if admin_user and target_user_id == admin_user["id"] and role not in {"admin", "owner"}:
                flash("You cannot remove your own admin role.", "warning")
                return redirect(url_for("account_admin"))

            old_role = (target_user["role"] or "").lower()
            if old_role == "admin" and role != "admin":
                admin_count = db.execute(
                    "SELECT COUNT(*) AS c FROM users WHERE lower(role) IN ('admin','owner')"
                ).fetchone()["c"]
                if admin_count <= 1:
                    flash("Cannot demote the last admin-level account.", "warning")
                    return redirect(url_for("account_admin"))

            if new_password:
                if enforce_password_policy() and role == "staff":
                    password_errors = password_strength_errors(new_password)
                    if password_errors:
                        flash("New password " + "; ".join(password_errors) + ".", "warning")
                        return redirect(url_for("account_admin"))

            db.execute(
                "UPDATE users SET full_name=?, username=?, role=? WHERE id=?",
                (full_name, username, role, target_user_id)
            )
            if new_password:
                db.execute(
                    "UPDATE users SET password_hash=?, password_must_change=0 WHERE id=?",
                    (generate_password_hash(new_password), target_user_id)
                )

            db.execute("DELETE FROM user_admin_days WHERE user_id=?", (target_user_id,))
            for day_name, programme in selected_values:
                db.execute("""
                    INSERT INTO user_admin_days (user_id, day_name, programme)
                    VALUES (?, ?, ?)
                    ON CONFLICT(user_id, day_name, programme) DO NOTHING
                """, (target_user_id, day_name, programme))

            db.commit()
            details = f"Updated {username} role={role} admin_days={len(selected_values)}"
            if new_password:
                details += " password=changed"
            log_audit(
                "admin_update_user",
                entity_type="user",
                entity_id=target_user_id,
                details=details,
            )
            flash(f"Updated account: {username}", "success")
            return redirect(url_for("account_admin"))

        flash("Unknown admin action.", "warning")
        return redirect(url_for("account_admin"))

    logs = db.execute("""
        SELECT created_at, username, action, entity_type, entity_id, status, details
        FROM audit_logs
        ORDER BY created_at DESC, id DESC
        LIMIT 500
    """).fetchall()
    user_rows = db.execute("""
        SELECT id, username, full_name, role, password_must_change, created_at
        FROM users
        ORDER BY username
    """).fetchall()
    day_rows = db.execute("""
        SELECT user_id, day_name, programme
        FROM user_admin_days
        ORDER BY user_id, day_name, programme
    """).fetchall()
    assignment_map = {}
    for row in day_rows:
        key = row["user_id"]
        if not admin_day_cell_allowed(row["day_name"], row["programme"]):
            continue
        assignment_map.setdefault(key, set()).add(f"{row['day_name']}|{row['programme']}")

    grouped_options = build_admin_day_grouped_options()

    return render_template(
        "account_admin.html",
        logs=logs,
        users=user_rows,
        assignment_map=assignment_map,
        grouped_options=grouped_options,
        email_enabled=email_enabled(),
        email_owner_only=email_owner_only_mode(),
        email_webhook_configured=bool(os.environ.get("TASTERIST_EMAIL_WEBHOOK_URL", "").strip()),
    )


def cron_token_valid():
    expected = os.environ.get("TASTERIST_CRON_TOKEN", "").strip()
    if not expected:
        return False
    provided = (
        request.headers.get("X-Tasterist-Cron-Token", "").strip()
        or request.args.get("token", "").strip()
    )
    if not provided:
        return False
    return secrets.compare_digest(provided, expected)


@app.route("/admin/email/weekly-report/send", methods=["POST"])
@owner_required
def send_weekly_report_now():
    try:
        result = send_weekly_admin_report(trigger="manual")
    except Exception as exc:
        flash(f"Weekly report email failed: {exc}", "danger")
        return redirect(url_for("account_admin"))
    if result.get("disabled"):
        flash("Email is disabled. Set TASTERIST_EMAIL_ENABLED=1 to send.", "warning")
        return redirect(url_for("account_admin"))
    flash(
        f"Weekly report email sent to {result['sent']} recipient(s): {', '.join(result['recipients']) or 'none'}",
        "success",
    )
    return redirect(url_for("account_admin"))


@app.route("/cron/weekly-admin-report", methods=["POST"])
def cron_weekly_admin_report():
    if not cron_token_valid():
        abort(403)
    try:
        result = send_weekly_admin_report(trigger="cron")
    except Exception as exc:
        return {
            "status": "error",
            "error": str(exc),
            "time": datetime.now().isoformat(timespec="seconds"),
        }, 500
    return {
        "status": "disabled" if result.get("disabled") else "ok",
        "sent": result["sent"],
        "recipients": result["recipients"],
        "owner_only": result["owner_only"],
        "time": datetime.now().isoformat(timespec="seconds"),
    }


@app.route("/admin/fix-pm-times", methods=["POST"])
@owner_required
def admin_fix_pm_times():
    include_preschool = request.form.get("include_preschool") == "1"
    force = request.form.get("force") == "1"
    try:
        result = run_pm_time_fix(force=force, include_preschool=include_preschool)
    except Exception as exc:
        flash(f"PM time fix failed: {exc}", "danger")
        return redirect(request.referrer or url_for("all_tasters"))

    log_audit(
        "manual_fix_pm_times_request",
        entity_type="system",
        entity_id="time-fix",
        details=(
            f"include_preschool={1 if include_preschool else 0} | "
            f"force={1 if force else 0} | "
            f"applied={1 if result.get('applied') else 0} | "
            f"tasters={result.get('tasters_updated', 0)} | "
            f"leavers={result.get('leavers_updated', 0)} | "
            f"class_start={result.get('class_start_updated', 0)} | "
            f"class_end={result.get('class_end_updated', 0)}"
        ),
    )
    if result.get("applied"):
        flash(
            "PM time fix applied: "
            f"tasters {result['tasters_updated']}, leavers {result['leavers_updated']}, "
            f"class start {result['class_start_updated']}, class end {result['class_end_updated']}.",
            "success",
        )
    else:
        flash("No PM time updates were required.", "warning")
    return redirect(request.referrer or url_for("all_tasters"))


@app.route("/tasters")
def all_tasters():
    def has_alien_text(value):
        text = str(value or "").strip()
        if not text:
            return False
        return bool(re.search(r"[^A-Za-z0-9\s\-\&\(\)\/\.,:'\"]", text))

    tasters_raw = query("""
        SELECT
            id, child, programme, location, session, class_name,
            taster_date, attended, bg, badge, notes
        FROM tasters
        ORDER BY taster_date DESC, child
    """)
    tasters = []
    for row in tasters_raw:
        item = dict(row)
        class_name = str(item.get("class_name") or "").strip()
        session_text = str(item.get("session") or "").strip()
        notes_text = str(item.get("notes") or "").strip()

        issues = []
        unknown_class = False
        if not class_name or class_name in {"?", "Unknown", "unknown"}:
            unknown_class = True
            issues.append("Unknown class label")
        if "?" in class_name:
            unknown_class = True
            issues.append(f"Alien class marker: {class_name}")
        if session_text and not re.search(r"\b\d{1,2}:\d{2}(?::\d{2})?\b", session_text):
            issues.append(f"Alien session text: {session_text}")
        if has_alien_text(class_name):
            issues.append(f"Alien class text: {class_name}")
        if has_alien_text(session_text):
            issues.append(f"Alien session chars: {session_text}")
        if has_alien_text(notes_text):
            issues.append(f"Alien notes chars: {notes_text}")

        if issues:
            unknown_class = True
        item["unknown_class"] = 1 if unknown_class else 0
        item["diagnostic_text"] = " | ".join(issues)
        tasters.append(item)

    leavers_raw = query("""
        SELECT child, programme, leave_month, leave_date, class_day, session, class_name, email, source
        FROM leavers
        ORDER BY leave_month DESC, leave_date DESC, child
    """)
    leavers = []
    for row in leavers_raw:
        item = dict(row)
        class_name = str(item.get("class_name") or "").strip()
        class_day = str(item.get("class_day") or "").strip()
        session_text = str(item.get("session") or "").strip()
        source_text = str(item.get("source") or "").strip()
        email_text = str(item.get("email") or "").strip()
        inferred_day = extract_day_name(class_day) or extract_day_name(session_text)

        issues = []
        unknown_class = False
        if not class_name or class_name in {"?", "Unknown", "unknown"}:
            unknown_class = True
            issues.append("Unknown class label")
        if not inferred_day:
            unknown_class = True
            if class_day:
                issues.append(f"Unknown class day: {class_day}")
            else:
                issues.append("Unknown class day")
        if class_day and not extract_day_name(class_day):
            issues.append(f"Alien class day text: {class_day}")
        if session_text and not re.search(r"\b\d{1,2}:\d{2}(?::\d{2})?\b", session_text):
            issues.append(f"Alien session text: {session_text}")
        if has_alien_text(class_name):
            issues.append(f"Alien class text: {class_name}")
        if has_alien_text(class_day):
            issues.append(f"Alien class-day chars: {class_day}")
        if has_alien_text(session_text):
            issues.append(f"Alien session chars: {session_text}")
        if has_alien_text(source_text):
            issues.append(f"Alien source chars: {source_text}")
        if has_alien_text(email_text):
            issues.append(f"Alien email chars: {email_text}")

        item["unknown_class"] = 1 if unknown_class else 0
        item["diagnostic_text"] = " | ".join(issues)
        item["resolved_day"] = inferred_day or "?"
        leavers.append(item)

    return render_template("all_tasters.html", tasters=tasters, leavers=leavers)

# ==========================================================
# ADD TASTER
# ==========================================================

@app.route("/add", methods=["GET", "POST"])
def add():
    programme = request.args.get("programme", "lockwood")
    db = get_db()

    if request.method == "POST":
        child = normalise_child_name(request.form["child"])
        taster_date = request.form["taster_date"]
        session_label = request.form["session"]
        class_name = request.form.get("class_name", "").strip()
        notes = request.form.get("notes", "").strip()
        sync_excel = False

        if not child or not taster_date:
            flash("Missing fields", "danger")
            return redirect(request.url)

        session_label = normalise_session_label(session_label)

        if not session_label:
            flash("Please choose a session", "danger")
            return redirect(request.url)

        taster_dt = _parse_iso_date(taster_date)
        if not taster_dt:
            flash("Invalid taster date.", "danger")
            return redirect(request.url)
        allowed, reason = _validate_programme_date_guardrails(db, programme, taster_dt)
        if not allowed:
            flash(reason, "warning")
            return redirect(request.url)

        db.execute("""
            INSERT INTO tasters
            (child, programme, location, session, class_name, taster_date, notes)
            VALUES (?,?,?,?,?,?,?)
        """, (
            child,
            programme,
            programme.title(),
            session_label,
            class_name,
            taster_date,
            notes,
        ))
        taster_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        db.commit()

        inserted = db.execute("SELECT * FROM tasters WHERE id=?", (taster_id,)).fetchone()
        actor_initials = user_initials((current_user() or {}).get("full_name", ""))
        sync_msg = "Excel sync skipped"
        sync_status = "ok"
        if sync_excel and inserted:
            ok, sync_msg = sync_taster_to_excel(inserted, mode="add", actor_initials=actor_initials)
            if not ok:
                flash(f"Taster saved in app, but Excel sync needs review: {sync_msg}", "warning")
                sync_status = "warn"

        log_audit(
            "add_taster",
            entity_type="taster",
            entity_id=taster_id,
            details=f"{child} | {programme} | {taster_date} {session_label} | excel_sync={sync_msg}",
            status=sync_status,
        )

        flash(f"Taster added for {child}", "success")
        return redirect(url_for("day_detail", date_str=taster_date, programme=programme))

    week_start_raw = request.args.get("week_start") or request.args.get("taster_date")
    if week_start_raw:
        try:
            anchor_date = datetime.fromisoformat(week_start_raw).date()
        except ValueError:
            anchor_date = date.today()
    else:
        anchor_date = date.today()

    week_start = anchor_date - timedelta(days=anchor_date.weekday())
    week_end = week_start + timedelta(days=6)
    week_days = build_week_schedule(programme, week_start)

    return render_template(
        "add.html",
        programme=programme,
        week_days=week_days,
        week_start=week_start,
        week_end=week_end,
        prev_week=(week_start - timedelta(days=7)).isoformat(),
        next_week=(week_start + timedelta(days=7)).isoformat(),
        today_str=date.today().isoformat()
    )


@app.route("/add/manual", methods=["GET", "POST"])
def add_manual_taster():
    programme = request.args.get("programme", "lockwood")
    db = get_db()

    if request.method == "POST":
        child = normalise_child_name(request.form.get("child", ""))
        taster_date = request.form.get("taster_date", "").strip()
        class_name = request.form.get("class_name", "").strip() or "Manual Session"
        session_label = normalise_session_label(request.form.get("session_label", "").strip())
        notes = request.form.get("notes", "").strip()
        sync_excel = False

        if not child or not taster_date or not session_label:
            flash("Name, date, and session label are required.", "danger")
            return redirect(request.url)

        taster_dt = _parse_iso_date(taster_date)
        if not taster_dt:
            flash("Invalid taster date.", "danger")
            return redirect(request.url)
        allowed, reason = _validate_programme_date_guardrails(db, programme, taster_dt)
        if not allowed:
            flash(reason, "warning")
            return redirect(request.url)

        db.execute("""
            INSERT INTO tasters
            (child, programme, location, session, class_name, taster_date, notes)
            VALUES (?,?,?,?,?,?,?)
        """, (
            child,
            programme,
            programme.title(),
            session_label,
            class_name,
            taster_date,
            notes,
        ))
        taster_id = db.execute("SELECT last_insert_rowid()").fetchone()[0]
        db.commit()

        inserted = db.execute("SELECT * FROM tasters WHERE id=?", (taster_id,)).fetchone()
        actor_initials = user_initials((current_user() or {}).get("full_name", ""))
        sync_msg = "Excel sync skipped"
        sync_status = "ok"
        if sync_excel and inserted:
            ok, sync_msg = sync_taster_to_excel(inserted, mode="add", actor_initials=actor_initials)
            if not ok:
                flash(f"Taster saved in app, but Excel sync needs review: {sync_msg}", "warning")
                sync_status = "warn"

        log_audit(
            "add_taster_manual",
            entity_type="taster",
            entity_id=taster_id,
            details=f"{child} | {programme} | {taster_date} {session_label} | excel_sync={sync_msg}",
            status=sync_status,
        )

        flash(f"Manual taster added for {child}", "success")
        return redirect(url_for("day_detail", date_str=taster_date, programme=programme))

    week_start_raw = request.args.get("week_start") or request.args.get("taster_date")
    if week_start_raw:
        try:
            anchor_date = datetime.fromisoformat(week_start_raw).date()
        except ValueError:
            anchor_date = date.today()
    else:
        anchor_date = date.today()
    week_start = anchor_date - timedelta(days=anchor_date.weekday())
    today_str = date.today().isoformat()
    return render_template(
        "add_manual_taster.html",
        programme=programme,
        week_start=week_start,
        today_str=today_str,
    )

# ==========================================================
# IMPORT / DEV
# ==========================================================

@app.route("/_routes")
def show_routes():
    if not is_env_true("TASTERIST_DEV_TOOLS_ENABLED", "0"):
        abort(404)
    if not is_owner_user(current_user()):
        abort(403)
    return "<br>".join(
        f"{rule.endpoint} → {rule.rule}"
        for rule in app.url_map.iter_rules()
    )


@app.route("/health")
def health():
    db = get_db()
    db.execute("SELECT 1").fetchone()
    return {
        "status": "ok",
        "time": datetime.now().isoformat(timespec="seconds"),
    }


@app.route("/cloud/preflight")
@admin_required
def cloud_preflight():
    sheets_path = Path(get_import_source_folder()).expanduser()
    db_taster_count = get_db().execute("SELECT COUNT(*) c FROM tasters").fetchone()["c"]
    postgres_url = DATABASE_URL

    def status(ok, ok_label="OK", bad_label="Needs Attention"):
        return ok_label if ok else bad_label

    checks = [
        {
            "name": "Runtime DB Backend",
            "path": DB_BACKEND,
            "ok": True,
            "detail": "Should match your production storage strategy.",
        },
        {
            "name": "Import Sheets Folder",
            "path": str(sheets_path),
            "ok": sheets_path.exists() and os.access(sheets_path, os.R_OK),
            "detail": "Folder must be readable in cloud for imports.",
        },
        {
            "name": ("Postgres Tasters Rows" if USING_POSTGRES else "SQLite Tasters Rows"),
            "path": str(db_taster_count),
            "ok": int(db_taster_count or 0) > 0,
            "detail": "Should be non-zero for production dashboard data.",
        },
        {
            "name": "DATABASE_URL",
            "path": "Configured" if postgres_url else "Missing",
            "ok": bool(postgres_url),
            "detail": "Required for Postgres runtime and backup/restore operations.",
        },
    ]

    if not USING_POSTGRES:
        db_path = Path(DB_FILE)
        db_parent = db_path.parent
        checks.insert(1, {
            "name": "Database Directory",
            "path": str(db_parent),
            "ok": db_parent.exists() and os.access(db_parent, os.W_OK),
            "detail": "Must exist and be writable by the app process.",
        })
        checks.insert(2, {
            "name": "Database File",
            "path": str(db_path),
            "ok": (db_path.exists() and os.access(db_path, os.W_OK)) or (not db_path.exists() and os.access(db_parent, os.W_OK)),
            "detail": "Either writable existing file or writable parent for first create.",
        })
    if postgres_url:
        pg_ok = False
        pg_path = "Unavailable"
        pg_detail = "Could not query Postgres taster count."
        try:
            import psycopg

            with psycopg.connect(postgres_url, connect_timeout=5) as pg_conn:
                with pg_conn.cursor() as cur:
                    cur.execute("SELECT COUNT(*) FROM tasters")
                    pg_count = int(cur.fetchone()[0] or 0)
            pg_path = str(pg_count)
            pg_ok = pg_count > 0
            pg_detail = "Mirror source should be non-zero."
        except Exception as exc:
            pg_path = f"Error: {exc}"
            pg_ok = False
            pg_detail = "Check DATABASE_URL value, network, and Postgres availability."

        checks.append({
            "name": "Postgres Direct Rows",
            "path": pg_path,
            "ok": pg_ok,
            "detail": pg_detail,
        })

    last_import = load_last_import_data() or {}
    import_ok = (last_import.get("exit_code", 0) == 0) if last_import else False
    import_warn = len(last_import.get("warnings", [])) if last_import else 0
    checks.append({
        "name": "Latest Import",
        "path": str(last_import.get("run_at") or "Not run"),
        "ok": import_ok and import_warn == 0,
        "detail": "Should be green before onboarding staff.",
    })

    overall_ok = all(c["ok"] for c in checks)
    return render_template(
        "cloud_preflight.html",
        checks=checks,
        status=status,
        overall_ok=overall_ok,
        last_import=last_import,
    )


@app.route("/cloud/restore-from-postgres", methods=["POST"])
@admin_required
def cloud_restore_from_postgres():
    if USING_POSTGRES:
        flash("Runtime already uses Postgres directly. Restore to SQLite is not required.", "info")
        return redirect(url_for("cloud_preflight"))

    postgres_url = os.environ.get("DATABASE_URL", "").strip()
    if not postgres_url:
        flash("DATABASE_URL is not set; cannot restore from Postgres.", "danger")
        return redirect(url_for("cloud_preflight"))

    restore_script = os.path.join(BASE_DIR, "scripts", "restore_sqlite_from_postgres.py")
    if not os.path.exists(restore_script):
        flash("Restore script is missing in this deploy.", "danger")
        return redirect(url_for("cloud_preflight"))

    cmd = [
        sys.executable,
        restore_script,
        "--sqlite", DB_FILE,
        "--postgres-url", postgres_url,
    ]
    timeout_raw = os.environ.get("TASTERIST_IMPORT_TIMEOUT_SEC", "120").strip()
    try:
        timeout_seconds = max(30, int(timeout_raw))
    except ValueError:
        timeout_seconds = 120

    try:
        close_request_db_if_open()
        result = subprocess.run(
            cmd, cwd=BASE_DIR, capture_output=True, text=True, timeout=timeout_seconds
        )
        log_parts = []
        if result.stdout:
            log_parts.append(result.stdout.strip())
        if result.stderr:
            log_parts.append(result.stderr.strip())
        log_text = "\n\n".join(part for part in log_parts if part).strip() or "(No output captured)"
        os.makedirs(os.path.dirname(RESTORE_LOG_FILE), exist_ok=True)
        with open(RESTORE_LOG_FILE, "w", encoding="utf-8") as f:
            f.write(log_text + "\n")
        if result.returncode == 0:
            taster_count = get_db().execute("SELECT COUNT(*) c FROM tasters").fetchone()["c"]
            flash(f"Restore from Postgres complete. Tasters now: {taster_count}.", "success")
        else:
            flash("Restore from Postgres failed. Check preflight and logs.", "warning")
    except subprocess.TimeoutExpired:
        flash(f"Restore timed out after {timeout_seconds}s.", "warning")
    except Exception as exc:
        flash(f"Restore failed: {exc}", "danger")

    return redirect(url_for("cloud_preflight"))


@app.route("/cloud/backup")
@admin_required
def cloud_backup():
    if USING_POSTGRES:
        flash("SQLite backup download is unavailable in Postgres runtime. Use Render Postgres backups.", "warning")
        return redirect(url_for("cloud_preflight"))

    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_name = f"tasterist-backup-{ts}.db"
    tmp_file = tempfile.NamedTemporaryFile(prefix="tasterist-backup-", suffix=".db", delete=False)
    tmp_file_path = tmp_file.name
    tmp_file.close()

    src = sqlite3.connect(DB_FILE, timeout=max(5, SQLITE_BUSY_TIMEOUT_MS // 1000))
    src.execute(f"PRAGMA busy_timeout = {SQLITE_BUSY_TIMEOUT_MS}")
    try:
        dst = sqlite3.connect(tmp_file_path, timeout=max(5, SQLITE_BUSY_TIMEOUT_MS // 1000))
        dst.execute(f"PRAGMA busy_timeout = {SQLITE_BUSY_TIMEOUT_MS}")
        try:
            src.backup(dst)
        finally:
            dst.close()
    finally:
        src.close()

    return send_file(
        tmp_file_path,
        as_attachment=True,
        download_name=backup_name,
        mimetype="application/octet-stream"
    )

@app.route("/import")
@admin_required
def import_page():
    db = get_db()
    db_path = Path(DB_FILE)
    import_source = get_import_source_folder()
    import_root = Path(import_source).expanduser()
    xlsx_count = 0
    if import_root.exists():
        xlsx_count = len([
            f for f in import_root.rglob("*.xlsx")
            if not f.name.startswith("~$")
        ])
    stats = {
        "user_count": db.execute("SELECT COUNT(*) c FROM users").fetchone()["c"],
        "taster_count": db.execute("SELECT COUNT(*) c FROM tasters").fetchone()["c"],
        "leaver_count": db.execute("SELECT COUNT(*) c FROM leavers").fetchone()["c"],
        "db_file": redact_database_url(DATABASE_URL) if USING_POSTGRES else str(db_path),
        "db_exists": True if USING_POSTGRES else db_path.exists(),
        "db_backend": DB_BACKEND,
        "import_source": str(import_root),
        "import_source_exists": import_root.exists(),
        "xlsx_count": xlsx_count,
    }
    return render_template(
        "import.html",
        last_import=load_last_import_data(),
        import_source=import_source,
        storage_stats=stats,
    )


@app.route("/import/upload", methods=["GET", "POST"])
@admin_required
def import_upload():
    if request.method == "GET":
        flash("Use the upload form on the Import page.", "warning")
        return redirect(url_for("import_page"))

    files = request.files.getlist("workbooks")
    if not files:
        flash("No files selected.", "warning")
        return redirect(url_for("import_page"))

    import_root = Path(get_import_source_folder()).expanduser()
    import_root.mkdir(parents=True, exist_ok=True)

    saved = 0
    skipped = 0
    for f in files:
        original_name = (f.filename or "").strip()
        if not original_name:
            skipped += 1
            continue
        safe_name = secure_filename(original_name)
        if not safe_name.lower().endswith(".xlsx"):
            skipped += 1
            continue
        m = re.search(r"(20\d{2})", safe_name)
        target_dir = import_root / m.group(1) if m else import_root
        target_dir.mkdir(parents=True, exist_ok=True)
        target_path = target_dir / safe_name
        f.save(target_path)
        saved += 1

    if saved == 0:
        flash("No valid .xlsx files uploaded.", "warning")
        return redirect(url_for("import_page"))

    flash(f"Uploaded {saved} workbook(s) to import storage.", "success")
    if skipped:
        flash(f"Skipped {skipped} file(s) (invalid or empty filename).", "warning")

    if request.form.get("run_after_upload") == "1":
        replace_requested = request.form.get("replace_all") == "1"
        replace = replace_requested and destructive_imports_enabled()
        if replace_requested and not replace:
            flash("Replace-all import is disabled in this environment.", "warning")
        rc, _ = run_import_process(trigger="upload", replace=replace)
        log_audit(
            "run_import",
            entity_type="system",
            entity_id="upload",
            details=f"Import trigger=upload replace={1 if replace else 0} rc={rc}",
            status="ok" if rc == 0 else "warn",
        )
        if rc == 0:
            flash("Import complete after upload.", "success")
        elif rc == 2 and USING_POSTGRES:
            flash("Upload complete. Import execution is disabled in Postgres runtime.", "warning")
        else:
            flash("Upload complete, but import finished with warnings/errors.", "warning")

    return redirect(url_for("import_page"))


@app.route("/import/run", methods=["GET", "POST"])
@admin_required
def import_run():
    if request.method == "GET":
        flash("Use the Run Full Import button on the Import page.", "warning")
        return redirect(url_for("import_page"))

    replace_requested = request.form.get("replace_all") == "1"
    replace = replace_requested and destructive_imports_enabled()
    if replace_requested and not replace:
        flash("Replace-all import is disabled in this environment.", "warning")
    rc, _ = run_import_process(trigger="manual", replace=replace)
    log_audit(
        "run_import",
        entity_type="system",
        entity_id="manual",
        details=f"Import trigger=manual replace={1 if replace else 0} rc={rc}",
        status="ok" if rc == 0 else "warn",
    )
    if rc == 0:
        flash("Import complete", "success")
    elif rc == 2 and USING_POSTGRES:
        flash("Import execution is disabled in Postgres runtime.", "warning")
    else:
        flash("Import finished with warnings/errors. Check the log.", "warning")
    return redirect(url_for("import_page"))


@app.route("/dev", methods=["GET", "POST"])
def dev_panel():
    if not is_env_true("TASTERIST_DEV_TOOLS_ENABLED", "0"):
        abort(404)
    if not is_owner_user(current_user()):
        abort(403)
    if request.method == "POST":
        db = get_db()
        db.execute("DELETE FROM tasters")
        db.execute("DELETE FROM leavers")
        db.commit()
        flash("Database cleared", "warning")
    return render_template("dev.html")

# ==========================================================
# BOOT
# ==========================================================

init_db()
maybe_restore_sqlite_from_postgres()
maybe_auto_fix_pm_times()
maybe_auto_fix_late_night_times()

try:
    _build_commit = os.environ.get("RENDER_GIT_COMMIT", "").strip()
    if _build_commit:
        print(f"🏷️ Build commit: {_build_commit[:7]}")
    _boot_db = open_db_connection()
    _boot_count = _boot_db.execute("SELECT COUNT(*) FROM tasters").fetchone()[0]
    _boot_db.close()
    _boot_target = "postgres" if USING_POSTGRES else DB_FILE
    print(f"🗄️ DB ready: backend={DB_BACKEND} target={_boot_target} | tasters={_boot_count}")
except Exception as exc:
    _boot_target = "postgres" if USING_POSTGRES else DB_FILE
    print(f"⚠️ DB startup check failed for backend={DB_BACKEND} target={_boot_target}: {exc}")

if __name__ == "__main__":
    app.run(debug=is_env_true("TASTERIST_DEBUG", "0"), port=8501)
