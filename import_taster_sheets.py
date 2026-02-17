#!/usr/bin/env python3
"""
Tasterist importer — CLEAN VERSION

PASS 1: detect structure
PASS 2: insert tasters
PASS 3: insert leavers (month-only)

• No CSV
• No duplicates
• Idempotent (safe to re-run)
"""

import sqlite3
import argparse
from pathlib import Path
from datetime import datetime
import re
import zipfile
from openpyxl import load_workbook

# --------------------------------------------------
# CONFIG
# --------------------------------------------------

DAYS = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]

MONTHS = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December"
]

# --------------------------------------------------
# HELPERS
# --------------------------------------------------

def normalise_programme(fname):
    f = fname.lower()
    # Support both "preschool" and "pre-school" filename variants.
    if "preschool" in f or "pre-school" in f:
        return "preschool"
    if "honley" in f:
        return "honley"
    return "lockwood"


def is_supported_workbook(fname):
    f = fname.lower()
    return "taster" in f and "leaver" in f

def normalise_time(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%H:%M")
    if hasattr(v, "hour"):  # datetime.time
        return f"{v.hour:02d}:{v.minute:02d}"
    return str(v).strip()

def normalise_cell_text(v):
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M")
    if hasattr(v, "hour"):  # datetime.time
        return f"{v.hour:02d}:{v.minute:02d}"
    return str(v).strip()

def truthy(v):
    if v is None:
        return 0
    if isinstance(v, bool):
        return 1 if v else 0
    s = str(v).strip().lower()
    if not s:
        return 0
    if "no" in s:
        return 0
    if "yes" in s:
        return 1
    return 1

def extract_year(fname):
    m = re.search(r"(20\d{2})", fname)
    return int(m.group(1)) if m else datetime.now().year

def looks_like_time(v):
    return isinstance(v, str) and re.match(r"\d{1,2}:\d{2}", v.strip())

def parse_date(val, month, year):
    if val is None:
        return None

    if isinstance(val, datetime):
        return val.date().isoformat()

    s = str(val).strip().lower()
    if not s:
        return None

    s = re.sub(r"(st|nd|rd|th)", "", s)
    s = re.sub(r"\bof\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()

    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            pass

    for fmt in ("%d/%m/%Y", "%d/%m", "%d-%b", "%d %b", "%d%b", "%d %B", "%d%B"):
        try:
            d = datetime.strptime(s, fmt)
            return d.replace(year=year).date().isoformat()
        except Exception:
            pass

    try:
        return datetime.strptime(
            f"{s} {month} {year}", "%d %B %Y"
        ).date().isoformat()
    except Exception:
        try:
            return datetime.strptime(
                f"{s} {month[:3]} {year}", "%d %b %Y"
            ).date().isoformat()
        except Exception:
            return None


# --------------------------------------------------
# IMPORT ONE WORKBOOK
# --------------------------------------------------

def find_name_columns(ws, max_scan_rows=25):
    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        cols = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == "name":
                cols.append(c)
        if cols:
            return r, cols
    return None, []


def find_section_rows(ws, marker):
    hits = []
    for r in range(1, ws.max_row + 1):
        row_has_marker = False
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() == marker.lower():
                row_has_marker = True
                break
        if row_has_marker:
            hits.append(r)
    return hits


def find_leaver_header_row(ws, start_row):
    scan_to = min(start_row + 15, ws.max_row)
    for r in range(start_row, scan_to + 1):
        name_cols = []
        has_leave_col = False
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            s = v.strip().lower()
            if s == "name":
                name_cols.append(c)
            if "leave" in s:
                has_leave_col = True
        if name_cols and has_leave_col:
            return r, name_cols
    return None, []


def import_excel(path, conn):
    programme = normalise_programme(path.name)
    year = extract_year(path.name)
    if programme == "honley":
        location = "Honley"
    elif programme == "preschool":
        location = "Preschool"
    else:
        location = "Lockwood"

    wb = load_workbook(path, data_only=True)
    cur = conn.cursor()

    def time_candidates(start_time):
        if not start_time or ":" not in start_time:
            return []
        out = [start_time]
        try:
            h, m = start_time.split(":", 1)
            hh = int(h)
            mm = m[:2]
        except ValueError:
            return out
        if 1 <= hh <= 11:
            out.append(f"{hh + 12:02d}:{mm}")
        return out

    def infer_class_name(programme_key, day_name, start_time, iso_date):
        if not start_time:
            return ""

        for candidate in time_candidates(start_time):
            row = cur.execute("""
                SELECT class_name
                FROM class_sessions
                WHERE programme=? AND session_date=? AND substr(start_time, 1, 5)=?
                ORDER BY class_name
                LIMIT 1
            """, (programme_key, iso_date, candidate)).fetchone()
            if row:
                return row[0] or ""

        weekday = day_name
        if not weekday:
            try:
                weekday = datetime.fromisoformat(iso_date).strftime("%A")
            except ValueError:
                weekday = ""

        for candidate in time_candidates(start_time):
            row = cur.execute("""
                SELECT class_name
                FROM class_sessions
                WHERE programme=? AND day=? AND substr(start_time, 1, 5)=?
                ORDER BY class_name
                LIMIT 1
            """, (programme_key, weekday, candidate)).fetchone()
            if row:
                return row[0] or ""
        return ""

    tasters_inserted = 0
    leavers_inserted = 0

    print(f"\n📘 FILE: {path.name} → {programme} {year}")

    for sheet_idx, ws in enumerate(wb.worksheets):
        month = MONTHS[sheet_idx]
        print(f"  • {month}")

        name_header_row, name_cols = find_name_columns(ws)
        if not name_cols:
            print("   ⚠️ No Name column found — skipping sheet")
            continue

        leaver_markers = find_section_rows(ws, "LEAVERS")
        taster_end_row = min(leaver_markers) - 1 if leaver_markers else ws.max_row

        sheet_default_date = datetime.strptime(
            f"1 {month} {year}", "%d %B %Y"
        ).date().isoformat()

        def header_text(col_idx):
            if col_idx < 1 or col_idx > ws.max_column:
                return ""
            v = ws.cell(name_header_row, col_idx).value
            return str(v).strip().lower() if v is not None else ""

        def find_col(name_col, fallback_offset, matcher):
            fallback = name_col + fallback_offset
            for c in range(name_col + 1, min(name_col + 11, ws.max_column + 1)):
                if matcher(header_text(c)):
                    return c
            return fallback

        column_map = {}
        for col in name_cols:
            day_col = col - 1
            date_col = find_col(
                col, 1,
                lambda t: "date" in t and ("taster" in t or "date of" in t)
            )
            attended_col = find_col(
                col, 2,
                lambda t: "attend" in t
            )
            club_fees_col = find_col(
                col, 3,
                lambda t: ("paid club fees" in t) or ("club fees" in t) or ("dd" in t and "paid" in t)
            )
            bg_col = find_col(
                col, 4,
                lambda t: ("paid bg" in t) or (t == "bg") or ("paid" in t and "bg" in t)
            )
            badge_col = find_col(
                col, 5,
                lambda t: ("added bg" in t) or ("badge" in t) or ("account" in t and "bg" in t)
            )
            notes_col = find_col(
                col, 6,
                lambda t: ("note" in t) or ("medical" in t)
            )
            column_map[col] = {
                "day_col": day_col,
                "date_col": date_col,
                "attended_col": attended_col,
                "club_fees_col": club_fees_col,
                "bg_col": bg_col,
                "badge_col": badge_col,
                "notes_col": notes_col,
            }

        # Per name-column state lets each day block carry its own day/time/date.
        block_state = {
            col: {"day": None, "time": None, "date": None}
            for col in name_cols
        }

        for r in range(name_header_row + 1, taster_end_row + 1):
            for col in name_cols:
                cols = column_map[col]
                day_col = cols["day_col"]
                date_col = cols["date_col"]
                attended_col = cols["attended_col"]
                club_fees_col = cols["club_fees_col"]
                bg_col = cols["bg_col"]
                badge_col = cols["badge_col"]
                notes_col = cols["notes_col"]

                if day_col >= 1:
                    day_or_time = ws.cell(r, day_col).value
                    if isinstance(day_or_time, str):
                        stripped = day_or_time.strip()
                        if stripped in DAYS:
                            block_state[col]["day"] = stripped
                            block_state[col]["time"] = None
                        elif looks_like_time(stripped):
                            block_state[col]["time"] = normalise_time(stripped)
                    elif hasattr(day_or_time, "hour"):
                        block_state[col]["time"] = normalise_time(day_or_time)

                parsed = parse_date(ws.cell(r, date_col).value, month, year)
                if parsed:
                    block_state[col]["date"] = parsed

                name_val = ws.cell(r, col).value
                if not isinstance(name_val, str):
                    continue

                name = name_val.strip()
                if not name or name.lower() == "name" or name.upper() == "LEAVERS":
                    continue

                if not block_state[col]["day"] and not block_state[col]["date"] and not parsed:
                    continue

                effective_date = parsed or block_state[col]["date"] or sheet_default_date
                session = block_state[col]["time"] or ""
                class_name = infer_class_name(
                    programme,
                    block_state[col]["day"],
                    block_state[col]["time"],
                    effective_date
                )

                note_val = ws.cell(r, notes_col).value if notes_col <= ws.max_column else ""

                cur.execute("""
                    INSERT OR IGNORE INTO tasters (
                        child, programme, location, session, class_name, taster_date,
                        attended, club_fees, bg, badge, notes
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    name,
                    programme,
                    location,
                    session,
                    class_name,
                    effective_date,
                    truthy(ws.cell(r, attended_col).value) if attended_col <= ws.max_column else 0,
                    truthy(ws.cell(r, club_fees_col).value) if club_fees_col <= ws.max_column else 0,
                    truthy(ws.cell(r, bg_col).value) if bg_col <= ws.max_column else 0,
                    truthy(ws.cell(r, badge_col).value) if badge_col <= ws.max_column else 0,
                    normalise_cell_text(note_val),
                ))

                if cur.rowcount == 1:
                    tasters_inserted += 1

        # -------- LEAVERS (structured section only) --------
        if leaver_markers:
            leaver_header_row, leaver_name_cols = find_leaver_header_row(ws, min(leaver_markers))
            if leaver_header_row and leaver_name_cols:
                seen_leavers = set()
                default_leave_month = f"{year}-{sheet_idx+1:02d}"

                for r in range(leaver_header_row + 1, ws.max_row + 1):
                    for col in leaver_name_cols:
                        name_val = ws.cell(r, col).value
                        if not isinstance(name_val, str):
                            continue

                        name = name_val.strip()
                        if not name or name.lower() == "name" or name.upper() == "LEAVERS":
                            continue

                        parsed_leave = parse_date(ws.cell(r, col + 1).value, month, year)
                        leave_month = parsed_leave[:7] if parsed_leave else default_leave_month
                        leave_date = parsed_leave or ""
                        sheet_day = ""
                        sheet_time = ""

                        for probe_col in range(max(1, col - 4), col):
                            probe_val = ws.cell(r, probe_col).value
                            probe_text = str(probe_val).strip() if probe_val is not None else ""
                            if probe_text in DAYS:
                                sheet_day = probe_text
                            if ":" in probe_text:
                                maybe_time = normalise_time(probe_text)
                                if maybe_time and ":" in maybe_time:
                                    sheet_time = maybe_time

                        if not sheet_day or not sheet_time:
                            for rr in range(r, max(leaver_header_row, r - 12), -1):
                                probe_val = ws.cell(rr, max(1, col - 1)).value
                                probe_text = str(probe_val).strip() if probe_val is not None else ""
                                if not sheet_day and probe_text in DAYS:
                                    sheet_day = probe_text
                                if not sheet_time and ":" in probe_text:
                                    maybe_time = normalise_time(probe_text)
                                    if maybe_time and ":" in maybe_time:
                                        sheet_time = maybe_time
                                if sheet_day and sheet_time:
                                    break

                        inferred_session = ""
                        inferred_class = ""
                        matched = cur.execute("""
                            SELECT session, class_name, taster_date
                            FROM tasters
                            WHERE lower(child)=lower(?) AND programme=?
                              AND substr(taster_date, 1, 7)=?
                            ORDER BY taster_date DESC
                            LIMIT 1
                        """, (name, programme, leave_month)).fetchone()
                        if not matched:
                            matched = cur.execute("""
                                SELECT session, class_name, taster_date
                                FROM tasters
                                WHERE lower(child)=lower(?) AND programme=?
                                ORDER BY taster_date DESC
                                LIMIT 1
                            """, (name, programme)).fetchone()
                        if matched:
                            inferred_session = matched[0] or ""
                            inferred_class = matched[1] or ""
                        if not inferred_session and (sheet_day or sheet_time):
                            inferred_session = " ".join([x for x in [sheet_day, sheet_time] if x]).strip()
                        dedupe_key = (name.lower(), leave_month)
                        if dedupe_key in seen_leavers:
                            continue
                        seen_leavers.add(dedupe_key)

                        cur.execute("""
                            INSERT OR IGNORE INTO leavers
                            (child, programme, leave_month, leave_date, session, class_name, source)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                        """, (
                            name,
                            programme,
                            leave_month,
                            leave_date,
                            inferred_session,
                            inferred_class,
                            path.name
                        ))

                        if cur.rowcount == 1:
                            leavers_inserted += 1

    print(f"   ✔ Tasters: {tasters_inserted}")
    print(f"   ✔ Leavers: {leavers_inserted}")
    return tasters_inserted, leavers_inserted


# --------------------------------------------------
# MAIN
# --------------------------------------------------

def main():
    p = argparse.ArgumentParser()
    p.add_argument("--folder", required=True)
    p.add_argument("--fallback-folder")
    p.add_argument("--db", default="tasterist.db")
    p.add_argument("--apply", action="store_true")
    args = p.parse_args()

    conn = sqlite3.connect(args.db)
    taster_cols = {r[1] for r in conn.execute("PRAGMA table_info(tasters)")}
    migrated = False
    if "class_name" not in taster_cols:
        conn.execute("ALTER TABLE tasters ADD COLUMN class_name TEXT DEFAULT ''")
        migrated = True
    if "club_fees" not in taster_cols:
        conn.execute("ALTER TABLE tasters ADD COLUMN club_fees INTEGER DEFAULT 0")
        migrated = True
    if migrated:
        conn.commit()

    if args.apply:
        print("\n🔥 Clearing tables")
        conn.execute("DELETE FROM tasters")
        conn.execute("DELETE FROM leavers")
        conn.commit()

    total_t = total_l = 0

    root = Path(args.folder).expanduser().resolve()

    if not root.exists():
        raise SystemExit(f"❌ Folder not found: {root}")

    print(f"\n📂 Importing from OneDrive path:")
    print(f"   {root}\n")

    fallback_lookup = {}
    if args.fallback_folder:
        fallback_root = Path(args.fallback_folder).expanduser().resolve()
        if fallback_root.exists():
            for fb in sorted(fallback_root.rglob("*.xlsx")):
                if fb.name.startswith("~$"):
                    continue
                fallback_lookup.setdefault(fb.name.lower(), fb)

    for file in sorted(root.rglob("*.xlsx")):
        # skip temp / lock files
        if file.name.startswith("~$"):
            continue
        if not is_supported_workbook(file.name):
            continue

        try:
            # OneDrive placeholder/unavailable files can fail with timeout or bad zip.
            import_path = file
            if not zipfile.is_zipfile(file):
                fallback = fallback_lookup.get(file.name.lower())
                if fallback and zipfile.is_zipfile(fallback):
                    print(f"ℹ️ Using local fallback: {file.name}")
                    import_path = fallback
                else:
                    print(f"⚠️ SKIP (invalid or not fully downloaded .xlsx): {file}")
                    continue

            t, l = import_excel(import_path, conn)
            total_t += t
            total_l += l
        except (TimeoutError, OSError, zipfile.BadZipFile) as exc:
            print(f"⚠️ SKIP (unreadable workbook): {file}")
            print(f"   ↳ {exc.__class__.__name__}: {exc}")
            continue
        except Exception as exc:
            print(f"⚠️ SKIP (unexpected import error): {file}")
            print(f"   ↳ {exc.__class__.__name__}: {exc}")
            continue


    conn.commit()
    conn.close()

    print("\n🎉 IMPORT COMPLETE")
    print(f"   ✔ Tasters: {total_t}")
    print(f"   ✔ Leavers: {total_l}")

if __name__ == "__main__":
    main()
